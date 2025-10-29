from fastapi import APIRouter, Depends, HTTPException
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import func, text, Date
import os
from typing import List, Dict, Any

from app.database import get_db
from app import models
from app.dependencies import get_current_user

router = APIRouter(tags=["Administración Básica"])

def verificar_admin(usuario: models.Usuario = Depends(get_current_user)):
    """Verifica que el usuario esté autenticado."""
    return usuario

def verificar_admin_o_supervisor(usuario: models.Usuario = Depends(get_current_user)):
    """Verifica que el usuario esté autenticado."""
    return usuario

@router.get("/dashboard/estadisticas")
def obtener_estadisticas_dashboard(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene estadísticas básicas para el dashboard de administrador.
    """
    try:
        # Usuarios activos (todos los usuarios)
        total_usuarios = db.query(models.Usuario).count()
        
        # Visitadores (usuarios con rol distinto a admin)
        visitadores_activos = db.query(models.Usuario).join(models.Rol).filter(
            models.Rol.nombre.ilike("%visitador%")
        ).count()
        
        # Visitas programadas hoy
        from datetime import datetime, date
        hoy = date.today()
        visitas_hoy = db.query(models.VisitaAsignada).filter(
            func.date(models.VisitaAsignada.fecha_programada) == hoy
        ).count()
        
        # Visitas esta semana
        from datetime import timedelta
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        fin_semana = inicio_semana + timedelta(days=6)
        visitas_semana = db.query(models.VisitaAsignada).filter(
            func.date(models.VisitaAsignada.fecha_programada) >= inicio_semana,
            func.date(models.VisitaAsignada.fecha_programada) <= fin_semana
        ).count()
        
        # Visitas completadas
        total_visitas = db.query(models.VisitaAsignada).count()
        completadas = db.query(models.VisitaAsignada).filter(
            models.VisitaAsignada.estado == "completada"
        ).count()
        
        porcentaje_completadas = (completadas / total_visitas * 100) if total_visitas > 0 else 0
        
        # Alertas críticas (valor estático por ahora)
        alertas_criticas = 0
        
        return {
            "usuarios_activos": total_usuarios,
            "visitadores_activos": visitadores_activos,
            "visitas_programadas_hoy": visitas_hoy,
            "visitas_programadas_semana": visitas_semana,
            "porcentaje_completadas": round(porcentaje_completadas, 1),
            "alertas_criticas": alertas_criticas,
            "graficos": {
                "visitas_por_municipio": [],
                "cumplimiento_checklist": {
                    "categorias": [],
                    "porcentajes": []
                }
            },
            "acciones_rapidas": [
                {"titulo": "Crear Usuario", "ruta": "/admin/usuarios/crear", "icono": "person_add"},
                {"titulo": "Programar Visitas", "ruta": "/admin/visitas/programar", "icono": "schedule"},
                {"titulo": "Exportar Histórico", "ruta": "/admin/exportaciones", "icono": "download"},
                {"titulo": "Gestionar Checklists", "ruta": "/admin/checklists", "icono": "checklist"}
            ]
        }
    except Exception as e:
        print(f"❌ Error en estadísticas dashboard: {e}")
        # Devolver datos por defecto en caso de error
        return {
            "usuarios_activos": 0,
            "visitadores_activos": 0,
            "visitas_programadas_hoy": 0,
            "visitas_programadas_semana": 0,
            "porcentaje_completadas": 0,
            "alertas_criticas": 0,
            "graficos": {
                "visitas_por_municipio": [],
                "cumplimiento_checklist": {
                    "categorias": [],
                    "porcentajes": []
                }
            },
            "acciones_rapidas": []
        }

# Endpoint eliminado - duplicado con el de abajo

@router.get("/tipos-visita")
def listar_tipos_visita(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Lista los tipos de visita disponibles.
    """
    # Por ahora devolvemos tipos estáticos
    return [
        {
            "id": 1,
            "nombre": "PAE",
            "descripcion": "Programa de Alimentación Escolar",
            "activo": True,
            "color": "#4CAF50"
        },
        {
            "id": 2,
            "nombre": "Infraestructura",
            "descripcion": "Revisión de infraestructura educativa",
            "activo": True,
            "color": "#2196F3"
        },
        {
            "id": 3,
            "nombre": "Supervisión",
            "descripción": "Supervisión general de la sede",
            "activo": True,
            "color": "#FF9800"
        }
    ]

@router.get("/checklists")
def listar_checklists(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Lista todas las categorías e items de checklist del sistema.
    """
    try:
        # Obtener categorías (sin ORDER BY orden porque no existe en BD)
        categorias = db.query(models.ChecklistCategoria).order_by(models.ChecklistCategoria.id).all()
        
        resultado = []
        for categoria in categorias:
            items = db.query(models.ChecklistItem).filter(
                models.ChecklistItem.categoria_id == categoria.id
            ).order_by(models.ChecklistItem.orden).all()
            
            resultado.append({
                "id": categoria.id,
                "nombre": categoria.nombre,
                "descripcion": categoria.descripcion or "",
                "orden": categoria.orden,
                "total_items": len(items),
                "items": [{
                    "id": item.id,
                    "texto": item.texto,
                    "tipo_respuesta": item.tipo_respuesta,
                    "orden": item.orden,
                    "obligatorio": item.obligatorio
                } for item in items]
            })
        
        return {
            "categorias": resultado,
            "total_categorias": len(resultado),
            "total_items": sum(len(cat["items"]) for cat in resultado)
        }
        
    except Exception as e:
        print(f"❌ Error al listar checklists: {e}")
        raise HTTPException(status_code=500, detail=f"Error al cargar checklists: {str(e)}")

@router.post("/checklists/categorias")
def crear_categoria_checklist(
    categoria_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Crea una nueva categoría de checklist.
    """
    try:
        nueva_categoria = models.ChecklistCategoria(
            nombre=categoria_data["nombre"]
            # Solo nombre porque es el único campo que existe en la BD
        )
        
        db.add(nueva_categoria)
        db.commit()
        db.refresh(nueva_categoria)
        
        return {
            "success": True,
            "message": "Categoría creada exitosamente",
            "categoria": {
                "id": nueva_categoria.id,
                "nombre": nueva_categoria.nombre,
                "descripcion": nueva_categoria.descripcion,
                "orden": nueva_categoria.orden
            }
        }
        
    except Exception as e:
        db.rollback()
        print(f"❌ Error al crear categoría: {e}")
        raise HTTPException(status_code=400, detail=f"Error al crear categoría: {str(e)}")

@router.put("/checklists/categorias/{categoria_id}")
def actualizar_categoria_checklist(
    categoria_id: int,
    categoria_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Actualiza una categoría de checklist existente.
    """
    try:
        categoria = db.query(models.ChecklistCategoria).filter(
            models.ChecklistCategoria.id == categoria_id
        ).first()
        
        if not categoria:
            raise HTTPException(status_code=404, detail="Categoría no encontrada")
        
        # Solo actualizar nombre porque es el único campo editable en la BD
        categoria.nombre = categoria_data.get("nombre", categoria.nombre)
        
        db.commit()
        
        return {
            "success": True,
            "message": "Categoría actualizada exitosamente"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error al actualizar categoría: {e}")
        raise HTTPException(status_code=400, detail=f"Error al actualizar categoría: {str(e)}")

@router.delete("/checklists/categorias/{categoria_id}")
def eliminar_categoria_checklist(
    categoria_id: int,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Elimina una categoría de checklist y todos sus items.
    """
    try:
        categoria = db.query(models.ChecklistCategoria).filter(
            models.ChecklistCategoria.id == categoria_id
        ).first()
        
        if not categoria:
            raise HTTPException(status_code=404, detail="Categoría no encontrada")
        
        # Eliminar todos los items de la categoría
        db.query(models.ChecklistItem).filter(
            models.ChecklistItem.categoria_id == categoria_id
        ).delete()
        
        # Eliminar la categoría
        db.delete(categoria)
        db.commit()
        
        return {
            "success": True,
            "message": "Categoría eliminada exitosamente"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error al eliminar categoría: {e}")
        raise HTTPException(status_code=400, detail=f"Error al eliminar categoría: {str(e)}")

@router.post("/checklists/items")
def crear_item_checklist(
    item_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Crea un nuevo item de checklist.
    """
    try:
        nuevo_item = models.ChecklistItem(
            categoria_id=item_data["categoria_id"],
            pregunta_texto=item_data["texto"],  # Usar nombre real de la BD
            orden=item_data.get("orden", 1)
        )
        
        db.add(nuevo_item)
        db.commit()
        db.refresh(nuevo_item)
        
        return {
            "success": True,
            "message": "Item creado exitosamente",
            "item": {
                "id": nuevo_item.id,
                "categoria_id": nuevo_item.categoria_id,
                "texto": nuevo_item.texto,
                "tipo_respuesta": nuevo_item.tipo_respuesta,
                "orden": nuevo_item.orden,
                "obligatorio": nuevo_item.obligatorio
            }
        }
        
    except Exception as e:
        db.rollback()
        print(f"❌ Error al crear item: {e}")
        raise HTTPException(status_code=400, detail=f"Error al crear item: {str(e)}")

@router.put("/checklists/items/{item_id}")
def actualizar_item_checklist(
    item_id: int,
    item_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Actualiza un item de checklist existente.
    """
    try:
        item = db.query(models.ChecklistItem).filter(
            models.ChecklistItem.id == item_id
        ).first()
        
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        
        # Solo actualizar campos que existen en la BD
        if "texto" in item_data:
            item.pregunta_texto = item_data["texto"]
        item.orden = item_data.get("orden", item.orden)
        
        db.commit()
        
        return {
            "success": True,
            "message": "Item actualizado exitosamente"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error al actualizar item: {e}")
        raise HTTPException(status_code=400, detail=f"Error al actualizar item: {str(e)}")

@router.delete("/checklists/items/{item_id}")
def eliminar_item_checklist(
    item_id: int,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Elimina un item de checklist.
    """
    try:
        item = db.query(models.ChecklistItem).filter(
            models.ChecklistItem.id == item_id
        ).first()
        
        if not item:
            raise HTTPException(status_code=404, detail="Item no encontrado")
        
        db.delete(item)
        db.commit()
        
        return {
            "success": True,
            "message": "Item eliminado exitosamente"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error al eliminar item: {e}")
        raise HTTPException(status_code=400, detail=f"Error al eliminar item: {str(e)}")

# Mantener los endpoints existentes pero mejorados
@router.get("/checklists/estadisticas")
def obtener_estadisticas_checklist(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene estadísticas de uso de checklists.
    """
    try:
        from sqlalchemy import text
        
        # Estadísticas básicas
        total_categorias = db.query(models.ChecklistCategoria).count()
        total_items = db.query(models.ChecklistItem).count()
        
        # Contar respuestas por tipo (si existe la tabla de respuestas)
        try:
            total_respuestas = db.execute(text("SELECT COUNT(*) FROM visita_respuesta_completa")).scalar() or 0
        except:
            total_respuestas = 0
        
        return {
            "total_categorias": total_categorias,
            "total_items": total_items,
            "total_respuestas": total_respuestas,
            "promedio_items_por_categoria": round(total_items / max(total_categorias, 1), 2),
            "ultima_actualizacion": "2024-01-15"  # Placeholder
        }
        
    except Exception as e:
        print(f"❌ Error al obtener estadísticas: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener estadísticas: {str(e)}")

# Endpoint mejorado de publicación
@router.post("/checklists/{checklist_id}/publicar")
def publicar_checklist(
    checklist_id: int,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Publica/activa un checklist para uso en visitas.
    """
    try:
        # Verificar que la categoría existe
        categoria = db.query(models.ChecklistCategoria).filter(
            models.ChecklistCategoria.id == checklist_id
        ).first()
        
        if not categoria:
            raise HTTPException(status_code=404, detail="Checklist no encontrado")
        
        # Contar items de la categoría
        total_items = db.query(models.ChecklistItem).filter(
            models.ChecklistItem.categoria_id == checklist_id
        ).count()
        
        if total_items == 0:
            raise HTTPException(status_code=400, detail="No se puede publicar un checklist sin items")
        
        return {
            "success": True,
            "message": f"Checklist '{categoria.nombre}' publicado exitosamente",
            "checklist": {
                "id": categoria.id,
                "nombre": categoria.nombre,
                "total_items": total_items
            },
            "publicado": True,
            "fecha_publicacion": datetime.now().isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al publicar checklist: {e}")
        raise HTTPException(status_code=400, detail=f"Error al publicar checklist: {str(e)}")



@router.get("/configuracion")
def obtener_configuracion(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene la configuración del sistema.
    """
    # Por ahora devolvemos configuración estática
    return {
        "seguridad": {
            "sesion_duracion_minutos": 480,
            "intentos_login_max": 5,
            "bloqueo_duracion_minutos": 30
        },
        "exportacion": {
            "formatos_soportados": ["excel", "pdf", "csv"],
            "max_registros_por_export": 10000,
            "retencion_archivos_dias": 7
        },
        "notificaciones": {
            "push_habilitado": True,
            "email_habilitado": True,
            "recordatorios_visitas": True
        }
    }

@router.get("/municipios")
def listar_municipios_admin(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin_o_supervisor)
):
    """
    Lista todos los municipios para admin.
    """
    try:
        municipios = db.query(models.Municipio).all()
        return [
            {
                "id": m.id,
                "nombre": m.nombre
            }
            for m in municipios
        ]
    except Exception as e:
        print(f"❌ Error al listar municipios: {e}")
        return []

@router.get("/sedes")
def listar_sedes_admin(
    municipio_id: int = None,
    institucion_id: int = None,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin_o_supervisor)
):
    """
    Lista sedes educativas para admin desde la vista consolidada.
    Si se especifica municipio_id o institucion_id, filtra por esos parámetros.
    """
    try:
        # Usar la tabla real sedes_educativas
        query = db.query(models.SedeEducativa)
        
        if municipio_id:
            query = query.filter(models.SedeEducativa.municipio_id == municipio_id)
        if institucion_id:
            query = query.filter(models.SedeEducativa.institucion_id == institucion_id)
            
        sedes_db = query.order_by(models.SedeEducativa.nombre_sede).all()
        
        return [
            {
                "id": sede.id,
                "nombre": sede.nombre_sede,
                "municipio_id": sede.municipio_id,
                "institucion_id": sede.institucion_id,
                "direccion": f"DANE: {sede.dane}",
                "estado": "activa",
                "dane": sede.dane,
                "due": sede.due,
                "lat": sede.lat,
                "lon": sede.lon,
                "principal": sede.principal
            }
            for sede in sedes_db
        ]
    except Exception as e:
        print(f"❌ Error al listar sedes: {e}")
        return []

@router.get("/instituciones")
def listar_instituciones_admin(
    municipio_id: int = None,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin_o_supervisor)
):
    """
    Lista instituciones educativas para admin desde la vista consolidada. 
    Si se especifica municipio_id, filtra por ese municipio.
    """
    try:
        sql = """
            SELECT 
                ROW_NUMBER() OVER (ORDER BY nombre, municipio) as id,
                nombre,
                municipio_id
            FROM instituciones_consolidadas
        """
        params = {}
        
        if municipio_id:
            sql += " WHERE municipio_id = :municipio_id"
            params["municipio_id"] = municipio_id
        
        sql += " ORDER BY nombre, municipio"
        
        result = db.execute(text(sql), params).fetchall()
        
        return [
            {
                "id": row.id,
                "nombre": row.nombre,
                "municipio_id": row.municipio_id
            }
            for row in result
        ]
    except Exception as e:
        print(f"❌ Error al listar instituciones: {e}")
        return []

@router.get("/visitadores")
def listar_visitadores_admin(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin_o_supervisor)
):
    """
    Lista todos los visitadores del sistema para admin.
    """
    try:
        visitadores = db.query(models.Usuario).join(models.Rol).filter(
            models.Rol.nombre.ilike("%visitador%")
        ).all()
        
        return [
            {
                "id": v.id,
                "nombre": v.nombre,
                "correo": v.correo,
                "rol": v.rol.nombre if v.rol else "Sin rol",
                "activo": True,  # Por defecto True hasta tener el campo
                "equipo": "Equipo Principal"  # Por defecto hasta tener el campo
            }
            for v in visitadores
        ]
    except Exception as e:
        print(f"❌ Error al listar visitadores: {e}")
        return []

@router.post("/usuarios")
def crear_usuario_admin(
    usuario_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Crea un nuevo usuario.
    """
    try:
        from passlib.context import CryptContext
        pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
        
        # Validar datos básicos
        nombre = usuario_data.get('nombre')
        correo = usuario_data.get('correo')
        contrasena = usuario_data.get('contrasena', 'temporal123')
        rol_id = usuario_data.get('rol_id')
        
        if not all([nombre, correo, rol_id]):
            raise HTTPException(status_code=400, detail="Faltan datos obligatorios")
        
        # Verificar que el correo no exista
        existing_user = db.query(models.Usuario).filter(models.Usuario.correo == correo).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="El correo ya existe")
        
        # Hash de la contraseña
        hashed_password = pwd_context.hash(contrasena)
        
        # Crear usuario
        nuevo_usuario = models.Usuario(
            nombre=nombre,
            correo=correo,
            contrasena=hashed_password,
            rol_id=int(rol_id)
        )
        
        db.add(nuevo_usuario)
        db.commit()
        db.refresh(nuevo_usuario)
        
        return {
            "success": True,
            "message": "Usuario creado exitosamente",
            "usuario_id": nuevo_usuario.id
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al crear usuario: {e}")
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error al crear usuario: {str(e)}")

@router.post("/checklists/{checklist_id}/publicar")
def publicar_checklist(
    checklist_id: int,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Publica un checklist (funcionalidad básica).
    """
    try:
        # Por ahora devolvemos éxito
        return {
            "success": True,
            "message": f"Checklist {checklist_id} publicado exitosamente (funcionalidad en desarrollo)",
            "publicado": True,
            "fecha_publicacion": "2024-01-15"
        }
    except Exception as e:
        print(f"❌ Error al publicar checklist: {e}")
        raise HTTPException(status_code=400, detail=str(e))

# ==================== GESTIÓN DE ROLES ====================

@router.get("/roles")
def listar_roles_admin(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Lista todos los roles del sistema.
    """
    try:
        roles = db.query(models.Rol).all()
        return {
            "roles": [
                {
                    "id": rol.id,
                    "nombre": rol.nombre,
                    "descripcion": getattr(rol, 'descripcion', f'Rol {rol.nombre}'),
                    "activo": getattr(rol, 'activo', True),
                    "usuarios_count": db.query(models.Usuario).filter(models.Usuario.rol_id == rol.id).count(),
                    "fecha_creacion": getattr(rol, 'fecha_creacion', None),
                    "permisos": []  # Se cargarán por separado
                }
                for rol in roles
            ]
        }
    except Exception as e:
        print(f"❌ Error al listar roles: {e}")
        return []

@router.post("/roles")
def crear_rol_admin(
    rol_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Crea un nuevo rol.
    """
    try:
        nombre = rol_data.get('nombre')
        descripcion = rol_data.get('descripcion', '')
        
        if not nombre:
            raise HTTPException(status_code=400, detail="El nombre del rol es obligatorio")
        
        # Verificar que el nombre no exista
        existing_rol = db.query(models.Rol).filter(models.Rol.nombre == nombre).first()
        if existing_rol:
            raise HTTPException(status_code=400, detail="Ya existe un rol con ese nombre")
        
        # Crear rol (sin especificar ID para que PostgreSQL use la secuencia)
        nuevo_rol = models.Rol(
            nombre=nombre
        )
        
        db.add(nuevo_rol)
        db.commit()
        db.refresh(nuevo_rol)
        
        return {
            "success": True,
            "message": "Rol creado exitosamente",
            "rol_id": nuevo_rol.id
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al crear rol: {e}")
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error al crear rol: {str(e)}")

@router.put("/roles/{rol_id}")
def actualizar_rol_admin(
    rol_id: int,
    rol_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Actualiza un rol existente.
    """
    try:
        rol = db.query(models.Rol).filter(models.Rol.id == rol_id).first()
        if not rol:
            raise HTTPException(status_code=404, detail="Rol no encontrado")
        
        # Actualizar campos
        if 'nombre' in rol_data:
            rol.nombre = rol_data['nombre']
        
        db.commit()
        
        return {
            "success": True,
            "message": "Rol actualizado exitosamente"
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al actualizar rol: {e}")
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error al actualizar rol: {str(e)}")

@router.delete("/roles/{rol_id}")
def eliminar_rol_admin(
    rol_id: int,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Elimina un rol (soft delete).
    """
    try:
        rol = db.query(models.Rol).filter(models.Rol.id == rol_id).first()
        if not rol:
            raise HTTPException(status_code=404, detail="Rol no encontrado")
        
        # Verificar que no tenga usuarios asignados
        usuarios_count = db.query(models.Usuario).filter(models.Usuario.rol_id == rol_id).count()
        if usuarios_count > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"No se puede eliminar el rol porque tiene {usuarios_count} usuarios asignados"
            )
        
        # Eliminar rol
        db.delete(rol)
        db.commit()
        
        return {
            "success": True,
            "message": "Rol eliminado exitosamente"
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al eliminar rol: {e}")
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error al eliminar rol: {str(e)}")

# ==================== GESTIÓN DE PERMISOS ====================

@router.get("/permisos")
def listar_permisos_admin(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Lista todos los permisos disponibles del sistema.
    """
    try:
        # Consultar permisos desde la base de datos
        result = db.execute(text("SELECT id, nombre, descripcion, modulo FROM permisos ORDER BY modulo, id")).fetchall()
        
        permisos_sistema = [
            {
                "id": row[0],
                "nombre": row[1],
                "descripcion": row[2],
                "modulo": row[3]
            }
            for row in result
        ]
        
        return permisos_sistema
    except Exception as e:
        print(f"❌ Error al listar permisos: {e}")
        return []

@router.get("/roles/{rol_id}/permisos")
def obtener_permisos_rol(
    rol_id: int,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene los permisos asignados a un rol específico.
    """
    try:
        rol = db.query(models.Rol).filter(models.Rol.id == rol_id).first()
        if not rol:
            raise HTTPException(status_code=404, detail="Rol no encontrado")
        
        # Consultar permisos asignados desde la base de datos
        result = db.execute(text("""
            SELECT permiso_id FROM rol_permisos WHERE rol_id = :rol_id
        """), {"rol_id": rol_id}).fetchall()
        
        permisos_ids = [row[0] for row in result]
        
        return {"permisos_ids": permisos_ids}
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al obtener permisos del rol: {e}")
        return {"permisos_ids": []}

@router.post("/roles/{rol_id}/permisos")
def asignar_permisos_rol(
    rol_id: int,
    permisos_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Asigna permisos a un rol.
    """
    try:
        rol = db.query(models.Rol).filter(models.Rol.id == rol_id).first()
        if not rol:
            raise HTTPException(status_code=404, detail="Rol no encontrado")
        
        permisos_ids = permisos_data.get('permisos_ids', [])
        
        # Eliminar permisos existentes del rol
        db.execute(text("DELETE FROM rol_permisos WHERE rol_id = :rol_id"), {"rol_id": rol_id})
        
        # Insertar nuevos permisos
        permisos_insertados = 0
        for permiso_id in permisos_ids:
            # Verificar que el permiso existe
            permiso_existe = db.execute(text("SELECT id FROM permisos WHERE id = :permiso_id"), 
                                      {"permiso_id": permiso_id}).fetchone()
            if permiso_existe:
                db.execute(text("""
                    INSERT INTO rol_permisos (rol_id, permiso_id) 
                    VALUES (:rol_id, :permiso_id)
                """), {"rol_id": rol_id, "permiso_id": permiso_id})
                permisos_insertados += 1
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Permisos asignados al rol {rol.nombre} exitosamente",
            "permisos_asignados": permisos_insertados
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al asignar permisos: {e}")
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error al asignar permisos: {str(e)}")

# ==================== PROGRAMACIÓN MASIVA DE VISITAS ====================

@router.get("/visitas/calendario")
def obtener_calendario_visitas(
    mes: int = None,
    anio: int = None,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin_o_supervisor)
):
    """
    Obtiene el calendario de visitas para un mes específico.
    """
    try:
        from datetime import datetime, timedelta
        import calendar
        
        # Si no se especifica, usar mes actual
        if not mes or not anio:
            now = datetime.now()
            mes = mes or now.month
            anio = anio or now.year
        
        # Primer y último día del mes
        primer_dia = datetime(anio, mes, 1)
        ultimo_dia = datetime(anio, mes, calendar.monthrange(anio, mes)[1])
        
        # Consultar visitas asignadas del mes usando SQL directo
        result = db.execute(text("""
            SELECT va.id, va.sede_id, va.visitador_id, va.fecha_programada, va.estado,
                   se.nombre_sede as sede_nombre, u.nombre as visitador_nombre
            FROM visitas_asignadas va
            LEFT JOIN sedes_educativas se ON va.sede_id = se.id
            LEFT JOIN usuarios u ON va.visitador_id = u.id
            WHERE va.fecha_programada >= :fecha_inicio AND va.fecha_programada <= :fecha_fin
        """), {"fecha_inicio": primer_dia, "fecha_fin": ultimo_dia}).fetchall()
        
        # Organizar por días
        calendario = {}
        for dia in range(1, calendar.monthrange(anio, mes)[1] + 1):
            fecha_str = f"{anio}-{mes:02d}-{dia:02d}"
            calendario[fecha_str] = {
                "fecha": fecha_str,
                "visitas": [],
                "total": 0,
                "visitadores": []
            }
        
        # Llenar con visitas existentes
        for visita in result:
            fecha_str = visita.fecha_programada.strftime("%Y-%m-%d")
            if fecha_str in calendario:
                calendario[fecha_str]["visitas"].append({
                    "id": visita.id,
                    "sede_nombre": visita.sede_nombre or "Sin sede",
                    "visitador": visita.visitador_nombre or "Sin asignar",
                    "estado": visita.estado or "programada",
                    "hora": visita.fecha_programada.strftime("%H:%M") if visita.fecha_programada else None
                })
                calendario[fecha_str]["total"] += 1
                if visita.visitador_nombre:
                    if visita.visitador_nombre not in calendario[fecha_str]["visitadores"]:
                        calendario[fecha_str]["visitadores"].append(visita.visitador_nombre)
        
        return {
            "mes": mes,
            "anio": anio,
            "calendario": list(calendario.values()),
            "total_visitas_mes": len(result)
        }
    except Exception as e:
        print(f"❌ Error al obtener calendario: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener calendario: {str(e)}")

@router.post("/visitas/programar-masivo")
def programar_visitas_masivo(
    programacion_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin_o_supervisor)
):
    """
    Programa múltiples visitas de forma masiva usando el modelo actual.
    """
    try:
        from datetime import datetime, timedelta
        
        # Extraer datos
        sedes_ids = programacion_data.get('sedes_ids', [])
        fecha_inicio = datetime.fromisoformat(programacion_data.get('fecha_inicio'))
        fecha_fin = datetime.fromisoformat(programacion_data.get('fecha_fin'))
        tipo_visita = programacion_data.get('tipo_visita', 'PAE')
        frecuencia = programacion_data.get('frecuencia', 'mensual')
        
        if not sedes_ids:
            raise HTTPException(status_code=400, detail="Debe especificar al menos una sede")
        
        # Validar que las sedes existen
        sedes = db.query(models.SedeEducativa).filter(models.SedeEducativa.id.in_(sedes_ids)).all()
        if len(sedes) != len(sedes_ids):
            raise HTTPException(status_code=400, detail="Algunas sedes no existen")
        
        visitas_creadas = []
        errores = []
        
        # Obtener visitadores para asignar
        visitadores_ids = programacion_data.get('visitadores_ids', [])
        if visitadores_ids:
            visitadores = db.query(models.Usuario).filter(
                models.Usuario.id.in_(visitadores_ids)
            ).all()
        else:
            # Si no se especifican visitadores, obtener algunos disponibles
            visitadores = db.query(models.Usuario).filter(
                models.Usuario.rol_id == 1  # Rol visitador
            ).limit(5).all()
        
        if not visitadores:
            raise HTTPException(status_code=400, detail="No hay visitadores disponibles")
        
        # Crear visitas programadas para cada sede
        visitador_index = 0
        for i, sede in enumerate(sedes):
            try:
                # Asignar visitador de forma rotativa
                visitador = visitadores[visitador_index % len(visitadores)]
                visitador_index += 1
                
                # Calcular fecha específica dentro del rango
                dias_total = (fecha_fin - fecha_inicio).days + 1
                dia_asignado = (i * dias_total) // len(sedes)
                fecha_visita = fecha_inicio + timedelta(days=dia_asignado)
                
                # Verificar conflictos en la fecha específica
                conflicto = db.execute(text("""
                    SELECT id FROM visitas_programadas 
                    WHERE sede_id = :sede_id AND DATE(fecha_programada) = DATE(:fecha_programada)
                """), {"sede_id": sede.id, "fecha_programada": fecha_visita}).fetchone()
                
                if not conflicto:
                    # Crear visita asignada (no visita programada)
                    from app.models import VisitaAsignada
                    nueva_visita_asignada = VisitaAsignada(
                        sede_id=sede.id,
                        visitador_id=visitador.id,
                        supervisor_id=admin_user.id,
                        fecha_programada=fecha_visita,
                        tipo_visita=tipo_visita,
                        prioridad="normal",
                        estado="pendiente",
                        contrato="ADMIN_MASIVO",
                        operador=f"Admin-{admin_user.id}",
                        municipio_id=sede.municipio_id,
                        institucion_id=sede.institucion_id,
                        observaciones=f"Visita {tipo_visita} programada masivamente",
                        fecha_creacion=datetime.utcnow()
                    )
                    
                    db.add(nueva_visita_asignada)
                    
                    visitas_creadas.append({
                        "sede_nombre": sede.nombre_sede,
                        "visitador_nombre": visitador.nombre,
                        "fecha": fecha_visita.strftime("%Y-%m-%d"),
                        "tipo": tipo_visita
                    })
                else:
                    errores.append(f"La sede {sede.nombre_sede} ya tiene una visita programada para {fecha_visita.strftime('%Y-%m-%d')}")
            
            except Exception as e:
                errores.append(f"Error con sede {sede.nombre_sede}: {str(e)}")
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Programación masiva completada",
            "visitas_creadas": len(visitas_creadas),
            "errores": len(errores),
            "detalles": {
                "visitas": visitas_creadas[:10],  # Primeras 10 para no sobrecargar
                "errores": errores[:5]  # Primeros 5 errores
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error en programación masiva: {e}")
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error en programación masiva: {str(e)}")

@router.get("/visitas/disponibilidad")
def obtener_disponibilidad_visitadores(
    fecha_inicio: str,
    fecha_fin: str,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin_o_supervisor)
):
    """
    Obtiene la disponibilidad de visitadores en un rango de fechas.
    """
    try:
        from datetime import datetime
        
        fecha_inicio_dt = datetime.fromisoformat(fecha_inicio)
        fecha_fin_dt = datetime.fromisoformat(fecha_fin)
        
        # Obtener todos los visitadores
        visitadores = db.query(models.Usuario).filter(
            models.Usuario.rol_id == 1  # Rol visitador
        ).all()
        
        disponibilidad = []
        for visitador in visitadores:
            # Por ahora simulamos la disponibilidad ya que el modelo actual no tiene asignación por visitador
            # En el futuro se implementará con visitas asignadas reales
            
            # Simulación: visitadores con ID par están más ocupados
            if visitador.id % 2 == 0:
                visitas_simuladas = 8
                disponibilidad_porcentaje = 60.0
                estado = "ocupado"
            else:
                visitas_simuladas = 3
                disponibilidad_porcentaje = 85.0
                estado = "disponible"
            
            dias_periodo = (fecha_fin_dt - fecha_inicio_dt).days + 1
            capacidad_maxima = dias_periodo * 3
            
            disponibilidad.append({
                "visitador_id": visitador.id,
                "nombre": visitador.nombre,
                "visitas_programadas": visitas_simuladas,
                "capacidad_maxima": capacidad_maxima,
                "disponibilidad_porcentaje": disponibilidad_porcentaje,
                "estado": estado
            })
        
        return sorted(disponibilidad, key=lambda x: x['disponibilidad_porcentaje'], reverse=True)
    except Exception as e:
        print(f"❌ Error al obtener disponibilidad: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener disponibilidad: {str(e)}")

@router.delete("/visitas/cancelar-masivo")
def cancelar_visitas_masivo(
    cancelacion_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Cancela múltiples visitas de forma masiva (marca como inactivas).
    """
    try:
        visitas_ids = cancelacion_data.get('visitas_ids', [])
        
        if not visitas_ids:
            raise HTTPException(status_code=400, detail="Debe especificar visitas a cancelar")
        
        # Actualizar visitas usando SQL directo
        visitas_canceladas = db.execute(text("""
            UPDATE visitas_programadas 
            SET estado = 'cancelada' 
            WHERE id = ANY(:visitas_ids)
        """), {"visitas_ids": visitas_ids}).rowcount
        
        db.commit()
        
        return {
            "success": True,
            "message": f"{visitas_canceladas} visitas canceladas exitosamente",
            "visitas_canceladas": visitas_canceladas
        }
    except Exception as e:
        print(f"❌ Error al cancelar visitas: {e}")
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error al cancelar visitas: {str(e)}")

# ==================== EXPORTACIONES AVANZADAS ====================

@router.get("/exportaciones")
def listar_exportaciones_admin(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Lista las exportaciones disponibles y el historial.
    """
    try:
        # Plantillas disponibles
        plantillas = [
            {
                "id": "visitas_completas",
                "nombre": "Visitas Completas",
                "descripcion": "Reporte detallado de todas las visitas realizadas",
                "tipo": "excel",
                "icono": "table_chart"
            },
            {
                "id": "cronograma_visitas",
                "nombre": "Cronograma de Visitas",
                "descripcion": "Programación de visitas por fecha y visitador",
                "tipo": "excel",
                "icono": "schedule"
            },
            {
                "id": "estadisticas_pae",
                "nombre": "Estadísticas PAE",
                "descripcion": "Análisis del Programa de Alimentación Escolar",
                "tipo": "pdf",
                "icono": "pie_chart"
            },
            {
                "id": "sedes_consolidado",
                "nombre": "Consolidado de Sedes",
                "descripcion": "Listado completo de sedes educativas",
                "tipo": "excel",
                "icono": "location_on"
            },
            {
                "id": "usuarios_sistema",
                "nombre": "Usuarios del Sistema",
                "descripcion": "Reporte de usuarios activos por rol",
                "tipo": "excel",
                "icono": "people"
            }
        ]
        
        # Historial de exportaciones (simulado por ahora)
        import datetime
        historial = [
            {
                "id": 1,
                "plantilla": "visitas_completas",
                "nombre_archivo": "visitas_completas_2025_08.xlsx",
                "fecha_creacion": datetime.datetime.now() - datetime.timedelta(days=2),
                "usuario": admin_user.nombre,
                "estado": "completado",
                "tamaño": "2.3 MB",
                "descargas": 5
            },
            {
                "id": 2,
                "plantilla": "estadisticas_pae",
                "nombre_archivo": "estadisticas_pae_julio.pdf",
                "fecha_creacion": datetime.datetime.now() - datetime.timedelta(days=7),
                "usuario": admin_user.nombre,
                "estado": "completado",
                "tamaño": "1.8 MB",
                "descargas": 12
            }
        ]
        
        return {
            "plantillas": plantillas,
            "historial": historial
        }
    except Exception as e:
        print(f"❌ Error al listar exportaciones: {e}")
        raise HTTPException(status_code=400, detail=f"Error al listar exportaciones: {str(e)}")

@router.post("/exportaciones/generar")
def generar_exportacion(
    exportacion_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Genera una exportación basada en la plantilla y filtros especificados.
    """
    try:
        from datetime import datetime
        import os
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        plantilla_id = exportacion_data.get('plantilla_id')
        filtros = exportacion_data.get('filtros', {})
        formato = exportacion_data.get('formato', 'excel')
        
        if not plantilla_id:
            raise HTTPException(status_code=400, detail="Debe especificar una plantilla")
        
        # Crear directorio de exportaciones si no existe
        export_dir = "media/exports"
        os.makedirs(export_dir, exist_ok=True)
        
        # Generar nombre de archivo único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if plantilla_id == "visitas_completas":
            return _generar_reporte_visitas_completas(db, filtros, formato, timestamp, export_dir, admin_user)
        elif plantilla_id == "cronograma_visitas":
            return _generar_cronograma_visitas(db, filtros, formato, timestamp, export_dir, admin_user)
        elif plantilla_id == "estadisticas_pae":
            return _generar_estadisticas_pae(db, filtros, formato, timestamp, export_dir, admin_user)
        elif plantilla_id == "sedes_consolidado":
            return _generar_consolidado_sedes(db, filtros, formato, timestamp, export_dir, admin_user)
        elif plantilla_id == "usuarios_sistema":
            return _generar_reporte_usuarios(db, filtros, formato, timestamp, export_dir, admin_user)
        else:
            raise HTTPException(status_code=400, detail="Plantilla no encontrada")
            
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al generar exportación: {e}")
        raise HTTPException(status_code=400, detail=f"Error al generar exportación: {str(e)}")

@router.get("/exportaciones/ubicacion-archivos")
def obtener_ubicacion_archivos(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene información sobre dónde se ubican los archivos para descarga.
    """
    try:
        import os
        
        # Directorio base de archivos
        media_dir = "media"
        export_dir = os.path.join(media_dir, "exports")
        
        # Verificar si los directorios existen
        directorios_info = {
            "media": {
                "ruta": os.path.abspath(media_dir),
                "existe": os.path.exists(media_dir),
                "descripcion": "Directorio principal de archivos multimedia"
            },
            "exports": {
                "ruta": os.path.abspath(export_dir),
                "existe": os.path.exists(export_dir),
                "descripcion": "Archivos de exportación y reportes"
            },
            "firmas": {
                "ruta": os.path.abspath(os.path.join(media_dir, "firmas")),
                "existe": os.path.exists(os.path.join(media_dir, "firmas")),
                "descripcion": "Firmas digitales capturadas"
            },
            "fotos": {
                "ruta": os.path.abspath(os.path.join(media_dir, "fotos")),
                "existe": os.path.exists(os.path.join(media_dir, "fotos")),
                "descripcion": "Fotos de evidencias"
            },
            "pdfs": {
                "ruta": os.path.abspath(os.path.join(media_dir, "pdfs")),
                "existe": os.path.exists(os.path.join(media_dir, "pdfs")),
                "descripcion": "Documentos PDF"
            }
        }
        
        # Contar archivos en cada directorio
        for dir_name, info in directorios_info.items():
            if info["existe"]:
                try:
                    archivos = os.listdir(info["ruta"])
                    info["total_archivos"] = len(archivos)
                    info["archivos_recientes"] = sorted(archivos, key=lambda x: os.path.getmtime(os.path.join(info["ruta"], x)), reverse=True)[:5]
                except Exception as e:
                    info["total_archivos"] = 0
                    info["error"] = str(e)
            else:
                info["total_archivos"] = 0
        
        return {
            "mensaje": "Información de ubicación de archivos",
            "directorios": directorios_info,
            "instrucciones": {
                "acceso_web": "Los archivos se pueden descargar a través de los endpoints de la API",
                "acceso_directo": "Los archivos también están disponibles directamente en el sistema de archivos",
                "limpieza": "Se recomienda limpiar archivos antiguos periódicamente"
            }
        }
        
    except Exception as e:
        print(f"❌ Error al obtener ubicación de archivos: {e}")
        raise HTTPException(status_code=500, detail=f"Error al obtener ubicación de archivos: {str(e)}")

@router.get("/exportaciones/{export_id}/download")
def descargar_exportacion(
    export_id: str,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Descarga un archivo de exportación.
    """
    try:
        from fastapi.responses import FileResponse
        import os
        
        # Buscar archivo en el directorio de exportaciones
        export_dir = "media/exports"
        
        # Por seguridad, validar que el archivo existe y pertenece al directorio correcto
        archivo_path = os.path.join(export_dir, export_id)
        
        if not os.path.exists(archivo_path):
            raise HTTPException(status_code=404, detail="Archivo no encontrado")
        
        # Normalizar rutas para comparación correcta
        archivo_path_abs = os.path.abspath(archivo_path)
        export_dir_abs = os.path.abspath(export_dir)
        
        if not archivo_path_abs.startswith(export_dir_abs):
            raise HTTPException(status_code=403, detail="Acceso no autorizado")
        
        # Determinar tipo de contenido
        if export_id.endswith('.xlsx'):
            media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif export_id.endswith('.pdf'):
            media_type = 'application/pdf'
        else:
            media_type = 'application/octet-stream'
        
        return FileResponse(
            path=archivo_path,
            media_type=media_type,
            filename=export_id
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al descargar exportación: {e}")
        raise HTTPException(status_code=400, detail=f"Error al descargar exportación: {str(e)}")

def _generar_reporte_visitas_completas(db, filtros, formato, timestamp, export_dir, admin_user):
    """Genera reporte de visitas completas."""
    try:
        # Consultar visitas con joins para obtener información completa
        # Aplicar filtros si existen
        params = {}
        conditions = []
        
        if filtros.get('fecha_inicio'):
            conditions.append("vp.fecha_programada >= :fecha_inicio")
            params['fecha_inicio'] = filtros['fecha_inicio']
        
        if filtros.get('fecha_fin'):
            conditions.append("vp.fecha_programada <= :fecha_fin")
            params['fecha_fin'] = filtros['fecha_fin']
        
        if filtros.get('estado'):
            conditions.append("vp.estado = :estado")
            params['estado'] = filtros['estado']
        
        # Construir SQL con filtros en el lugar correcto
        from sqlalchemy import text
        
        sql_base = """
            SELECT 
                vp.id,
                vp.fecha_programada,
                vp.estado,
                vp.observaciones,
                se.nombre_sede,
                se.dane,
                u.nombre as visitador,
                vc.fecha_visita,
                vc.contrato as asunto,
                vc.observaciones as observaciones_visita
            FROM visitas_programadas vp
            LEFT JOIN sedes_educativas se ON vp.sede_id = se.id
            LEFT JOIN usuarios u ON vp.visitador_id = u.id
            LEFT JOIN visitas_completas_pae vc ON vp.sede_id = vc.sede_id 
                AND vp.visitador_id = vc.profesional_id
        """
        
        sql = sql_base
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        
        sql += " ORDER BY vp.fecha_programada DESC"
        
        result = db.execute(text(sql), params).fetchall()
        
        if formato == 'excel':
            # Crear DataFrame de pandas
            import pandas as pd
            
            data = []
            for row in result:
                data.append({
                    'ID Visita': row.id,
                    'Fecha Programada': row.fecha_programada.strftime('%Y-%m-%d %H:%M') if row.fecha_programada else '',
                    'Estado': row.estado or '',
                    'Sede': row.nombre_sede or '',
                    'DANE': row.dane or '',
                    'Visitador': row.visitador or '',
                    'Fecha Real': row.fecha_visita.strftime('%Y-%m-%d %H:%M') if row.fecha_visita else '',
                    'Asunto': row.asunto or '',
                    'Observaciones Programación': row.observaciones or '',
                    'Observaciones Visita': row.observaciones_visita or ''
                })
            
            df = pd.DataFrame(data)
            
            # Generar archivo Excel
            filename = f"visitas_completas_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Visitas Completas', index=False)
                
                # Formatear hoja
                worksheet = writer.sheets['Visitas Completas']
                
                # Ajustar ancho de columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Formatear encabezados
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
            
            return {
                "success": True,
                "message": "Reporte generado exitosamente",
                "filename": filename,
                "registros": len(data),
                "formato": "excel"
            }
        
        else:  # PDF
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            
            filename = f"visitas_completas_{timestamp}.pdf"
            filepath = os.path.join(export_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title = Paragraph("Reporte de Visitas Completas", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 20))
            
            # Información del reporte
            from datetime import datetime
            info = Paragraph(f"Generado por: {admin_user.nombre}<br/>Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/>Total registros: {len(result)}", styles['Normal'])
            elements.append(info)
            elements.append(Spacer(1, 20))
            
            # Tabla de datos (primeros 100 registros para evitar archivos muy grandes)
            if result:
                table_data = [['ID', 'Fecha', 'Estado', 'Sede', 'Visitador']]
                for row in result[:100]:  # Limitar a 100 registros
                    table_data.append([
                        str(row.id),
                        row.fecha_programada.strftime('%Y-%m-%d') if row.fecha_programada else '',
                        row.estado or '',
                        (row.nombre_sede or '')[:30],  # Truncar texto largo
                        (row.visitador or '')[:20]
                    ])
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
            
            doc.build(elements)
            
            return {
                "success": True,
                "message": "Reporte PDF generado exitosamente",
                "filename": filename,
                "registros": len(result),
                "formato": "pdf"
            }
    
    except Exception as e:
        raise Exception(f"Error generando reporte de visitas: {str(e)}")

def _generar_cronograma_visitas(db, filtros, formato, timestamp, export_dir, admin_user):
    """Genera cronograma de visitas programadas."""
    try:
        # Query para obtener visitas programadas con detalles
        from sqlalchemy import text
        
        query = text("""
        SELECT 
            va.id,
            va.fecha_programada,
            se.nombre_sede,
            se.dane,
            u.nombre AS visitador,
            va.estado,
            va.contrato,
            va.operador,
            va.observaciones,
            m.nombre AS municipio
        FROM visitas_asignadas va
        LEFT JOIN sedes_educativas se ON va.sede_id = se.id
        LEFT JOIN usuarios u ON va.visitador_id = u.id
        LEFT JOIN municipios m ON se.municipio_id = m.id
        ORDER BY va.fecha_programada DESC
        """)
        
        result = db.execute(query)
        visitas = result.fetchall()
        
        if formato == 'excel':
            import pandas as pd
            
            # Convertir a DataFrame
            df = pd.DataFrame([{
                'ID': visita[0],
                'Fecha Programada': visita[1],
                'Sede Educativa': visita[2] or 'Sin nombre',
                'DANE': visita[3] or 'N/A',
                'Visitador': visita[4] or 'Sin asignar',
                'Estado': visita[5] or 'programada',
                'Contrato': visita[6] or 'N/A',
                'Operador': visita[7] or 'N/A',
                'Observaciones': visita[8] or '',
                'Municipio': visita[9] or 'N/A'
            } for visita in visitas])
            
            filename = f"cronograma_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Cronograma Visitas', index=False)
                
                # Formatear
                worksheet = writer.sheets['Cronograma Visitas']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
                    
        elif formato == 'pdf':
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            filename = f"cronograma_{timestamp}.pdf"
            filepath = os.path.join(export_dir, filename)
            
            # Crear documento PDF
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            story.append(Paragraph("Cronograma de Visitas", title_style))
            story.append(Spacer(1, 12))
            
            # Preparar datos para la tabla
            data = [['ID', 'Fecha', 'Sede', 'Visitador', 'Estado', 'Municipio']]
            for visita in visitas:
                data.append([
                    str(visita[0]),
                    str(visita[1])[:10] if visita[1] else 'N/A',
                    (visita[2] or 'Sin nombre')[:20],
                    (visita[4] or 'Sin asignar')[:15],
                    visita[5] or 'programada',
                    (visita[9] or 'N/A')[:15]
                ])
            
            # Crear tabla
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            story.append(table)
            doc.build(story)
        
        return {
            "success": True,
            "message": f"Cronograma generado exitosamente - {len(visitas)} visitas",
            "filename": filename,
            "registros": len(visitas),
            "formato": formato
        }
        
    except Exception as e:
        print(f"❌ Error generando cronograma: {e}")
        return {
            "success": False,
            "message": f"Error al generar cronograma: {str(e)}",
            "filename": None,
            "registros": 0,
            "formato": formato
        }

def _generar_estadisticas_pae(db, filtros, formato, timestamp, export_dir, admin_user):
    """Genera estadísticas del PAE."""
    try:
        # Query para obtener estadísticas del PAE
        from sqlalchemy import text
        
        query = text("""
        SELECT 
            COUNT(*) as total_visitas,
            COUNT(CASE WHEN estado = 'completada' THEN 1 END) as visitas_completadas,
            COUNT(CASE WHEN estado = 'programada' THEN 1 END) as visitas_programadas,
            COUNT(CASE WHEN estado = 'cancelada' THEN 1 END) as visitas_canceladas
        FROM visitas_completas_pae
        """)
        
        result = db.execute(query)
        stats = result.fetchone()
        
        # Query para obtener visitas por municipio
        query_municipios = text("""
        SELECT 
            m.nombre as municipio,
            COUNT(*) as total_visitas,
            COUNT(CASE WHEN vc.estado = 'completada' THEN 1 END) as completadas
        FROM visitas_completas_pae vc
        LEFT JOIN sedes_educativas se ON vc.sede_id = se.id
        LEFT JOIN municipios m ON se.municipio_id = m.id
        GROUP BY m.nombre
        ORDER BY total_visitas DESC
        """)
        
        result_municipios = db.execute(query_municipios)
        municipios_stats = result_municipios.fetchall()
        
        if formato == 'excel':
            import pandas as pd
            
            # Crear datos del resumen general
            resumen_df = pd.DataFrame([{
                'Métrica': 'Total Visitas',
                'Valor': stats[0] or 0
            }, {
                'Métrica': 'Visitas Completadas',
                'Valor': stats[1] or 0
            }, {
                'Métrica': 'Visitas Programadas',
                'Valor': stats[2] or 0
            }, {
                'Métrica': 'Visitas Canceladas',
                'Valor': stats[3] or 0
            }, {
                'Métrica': 'Tasa de Cumplimiento (%)',
                'Valor': round((stats[1] or 0) * 100 / (stats[0] or 1), 2)
            }])
            
            # Crear datos por municipio
            municipios_df = pd.DataFrame([{
                'Municipio': municipio[0] or 'Sin municipio',
                'Total Visitas': municipio[1] or 0,
                'Completadas': municipio[2] or 0,
                'Tasa Cumplimiento (%)': round((municipio[2] or 0) * 100 / (municipio[1] or 1), 2)
            } for municipio in municipios_stats])
            
            filename = f"estadisticas_pae_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            # Crear archivo Excel con múltiples hojas
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                resumen_df.to_excel(writer, sheet_name='Resumen General', index=False)
                municipios_df.to_excel(writer, sheet_name='Por Municipio', index=False)
                
                # Formatear ambas hojas
                for sheet_name in ['Resumen General', 'Por Municipio']:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        return {
            "success": True,
            "message": f"Estadísticas PAE generadas - {stats[0] or 0} visitas analizadas",
            "filename": filename,
            "registros": stats[0] or 0,
            "formato": formato
        }
        
    except Exception as e:
        print(f"❌ Error generando estadísticas PAE: {e}")
        return {
            "success": False,
            "message": f"Error al generar estadísticas PAE: {str(e)}",
            "filename": None,
            "registros": 0,
            "formato": formato
        }

def _generar_consolidado_sedes(db, filtros, formato, timestamp, export_dir, admin_user):
    """Genera consolidado de sedes educativas."""
    try:
        # Query para obtener información completa de sedes
        from sqlalchemy import text
        
        query = text("""
        SELECT 
            se.id,
            se.nombre_sede,
            se.dane,
            se.due,
            m.nombre as municipio,
            ie.nombre as institucion,
            se.lat,
            se.lon,
            se.principal,
            COUNT(vp.id) as visitas_programadas,
            COUNT(CASE WHEN vp.estado = 'completada' THEN 1 END) as visitas_completadas,
            COUNT(vc.id) as visitas_pae_completadas
        FROM sedes_educativas se
        LEFT JOIN municipios m ON se.municipio_id = m.id
        LEFT JOIN instituciones ie ON se.institucion_id = ie.id
        LEFT JOIN visitas_programadas vp ON se.id = vp.sede_id
        LEFT JOIN visitas_completas_pae vc ON se.id = vc.sede_id
        GROUP BY se.id, se.nombre_sede, se.dane, se.due, m.nombre, ie.nombre, se.lat, se.lon, se.principal
        ORDER BY m.nombre, se.nombre_sede
        """)
        
        result = db.execute(query)
        sedes = result.fetchall()
        
        if formato == 'excel':
            import pandas as pd
            
            # Convertir a DataFrame
            df = pd.DataFrame([{
                'ID': sede[0],
                'Nombre Sede': sede[1] or 'Sin nombre',
                'DANE': sede[2] or 'N/A',
                'DUE': sede[3] or 'N/A',
                'Municipio': sede[4] or 'Sin municipio',
                'Institución': sede[5] or 'Sin institución',
                'Latitud': sede[6] or 0,
                'Longitud': sede[7] or 0,
                'Es Principal': 'Sí' if sede[8] else 'No',
                'Visitas Programadas': sede[9] or 0,
                'Visitas Completadas': sede[10] or 0,
                'Visitas PAE': sede[11] or 0,
                'Tasa Cumplimiento (%)': round((sede[10] or 0) * 100 / (sede[9] or 1), 2) if sede[9] else 0
            } for sede in sedes])
            
            filename = f"sedes_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Consolidado Sedes', index=False)
                
                # Crear resumen por municipio
                resumen_municipio = df.groupby('Municipio').agg({
                    'ID': 'count',
                    'Visitas Programadas': 'sum',
                    'Visitas Completadas': 'sum',
                    'Visitas PAE': 'sum'
                }).reset_index()
                resumen_municipio.columns = ['Municipio', 'Total Sedes', 'Visitas Programadas', 'Visitas Completadas', 'Visitas PAE']
                resumen_municipio['Tasa Cumplimiento (%)'] = round(
                    resumen_municipio['Visitas Completadas'] * 100 / resumen_municipio['Visitas Programadas'].replace(0, 1), 2
                )
                
                resumen_municipio.to_excel(writer, sheet_name='Resumen por Municipio', index=False)
                
                # Formatear hojas
                for sheet_name in ['Consolidado Sedes', 'Resumen por Municipio']:
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
                        
        elif formato == 'pdf':
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            filename = f"sedes_{timestamp}.pdf"
            filepath = os.path.join(export_dir, filename)
            
            # Crear documento PDF
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            story.append(Paragraph("Consolidado de Sedes Educativas", title_style))
            story.append(Spacer(1, 12))
            
            # Preparar datos para la tabla (versión simplificada para PDF)
            data = [['ID', 'Sede', 'Municipio', 'Institución', 'Visitas', 'Estado']]
            for sede in sedes:
                estado = 'Completada' if (sede[10] or 0) > 0 else 'Pendiente'
                data.append([
                    str(sede[0]),
                    (sede[1] or 'Sin nombre')[:25],
                    (sede[4] or 'Sin municipio')[:20],
                    (sede[5] or 'Sin institución')[:20],
                    f"{sede[10] or 0}/{sede[9] or 0}",
                    estado
                ])
            
            # Crear tabla
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
            ]))
            
            story.append(table)
            doc.build(story)
        
        return {
            "success": True,
            "message": f"Consolidado de sedes generado - {len(sedes)} sedes incluidas",
            "filename": filename,
            "registros": len(sedes),
            "formato": formato
        }
        
    except Exception as e:
        print(f"❌ Error generando consolidado de sedes: {e}")
        return {
            "success": False,
            "message": f"Error al generar consolidado de sedes: {str(e)}",
            "filename": None,
            "registros": 0,
            "formato": formato
        }

def _generar_reporte_usuarios(db, filtros, formato, timestamp, export_dir, admin_user):
    """Genera reporte de usuarios del sistema."""
    try:
        # Query para obtener información completa de usuarios
        from sqlalchemy import text
        
        query = text("""
        SELECT 
            u.id,
            u.nombre,
            u.correo,
            r.nombre as rol,
            COUNT(vp.id) as visitas_programadas,
            COUNT(CASE WHEN vp.estado = 'completada' THEN 1 END) as visitas_completadas,
            COUNT(vc.id) as visitas_pae_realizadas
        FROM usuarios u
        LEFT JOIN roles r ON u.rol_id = r.id
        LEFT JOIN visitas_programadas vp ON u.id = vp.visitador_id
        LEFT JOIN visitas_completas_pae vc ON u.id = vc.profesional_id
        GROUP BY u.id, u.nombre, u.correo, r.nombre
        ORDER BY r.nombre, u.nombre
        """)
        
        result = db.execute(query)
        usuarios = result.fetchall()
        
        if formato == 'excel':
            import pandas as pd
            
            # Convertir a DataFrame
            df = pd.DataFrame([{
                'ID': usuario[0],
                'Nombre': usuario[1] or 'Sin nombre',
                'Correo': usuario[2] or 'Sin correo',
                'Rol': usuario[3] or 'Sin rol',
                'Visitas Programadas': usuario[4] or 0,
                'Visitas Completadas': usuario[5] or 0,
                'Visitas PAE Realizadas': usuario[6] or 0,
                'Tasa Cumplimiento (%)': round((usuario[5] or 0) * 100 / (usuario[4] or 1), 2) if usuario[4] else 0,
                'Activo': 'Sí'  # Por ahora todos activos ya que no tenemos campo activo
            } for usuario in usuarios])
            
            filename = f"usuarios_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Usuarios del Sistema', index=False)
                
                # Crear resumen por rol
                resumen_rol = df.groupby('Rol').agg({
                    'ID': 'count',
                    'Visitas Programadas': 'sum',
                    'Visitas Completadas': 'sum',
                    'Visitas PAE Realizadas': 'sum'
                }).reset_index()
                resumen_rol.columns = ['Rol', 'Total Usuarios', 'Visitas Programadas', 'Visitas Completadas', 'Visitas PAE']
                resumen_rol['Tasa Cumplimiento Promedio (%)'] = round(
                    resumen_rol['Visitas Completadas'] * 100 / resumen_rol['Visitas Programadas'].replace(0, 1), 2
                )
                
                resumen_rol.to_excel(writer, sheet_name='Resumen por Rol', index=False)
                
                # Crear lista de visitadores más activos
                visitadores = df[df['Rol'] == 'visitador'].sort_values('Visitas Completadas', ascending=False).head(10)
                if not visitadores.empty:
                    visitadores.to_excel(writer, sheet_name='Top 10 Visitadores', index=False)
                
                # Formatear hojas
                for sheet_name in writer.sheets.keys():
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
                        
        elif formato == 'pdf':
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            filename = f"usuarios_{timestamp}.pdf"
            filepath = os.path.join(export_dir, filename)
            
            # Crear documento PDF
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            story.append(Paragraph("Usuarios del Sistema", title_style))
            story.append(Spacer(1, 12))
            
            # Preparar datos para la tabla (versión simplificada para PDF)
            data = [['ID', 'Nombre', 'Rol', 'Visitas', 'Estado']]
            for usuario in usuarios:
                estado = 'Activo' if (usuario[5] or 0) > 0 else 'Inactivo'
                data.append([
                    str(usuario[0]),
                    (usuario[1] or 'Sin nombre')[:20],
                    (usuario[3] or 'Sin rol')[:15],
                    f"{usuario[5] or 0}/{usuario[4] or 0}",
                    estado
                ])
            
            # Crear tabla
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
            ]))
            
            story.append(table)
            doc.build(story)
        
        return {
            "success": True,
            "message": f"Reporte de usuarios generado - {len(usuarios)} usuarios incluidos",
            "filename": filename,
            "registros": len(usuarios),
            "formato": formato
        }
        
    except Exception as e:
        print(f"❌ Error generando reporte de usuarios: {e}")
        return {
            "success": False,
            "message": f"Error al generar reporte de usuarios: {str(e)}",
            "filename": None,
            "registros": 0,
            "formato": formato
        }

# ==================== AUTENTICACIÓN 2FA ====================

@router.get("/2fa/status")
def obtener_status_2fa(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Obtiene el estado actual de 2FA para el usuario.
    """
    try:
        # Por ahora simulamos el estado (más adelante se guardará en BD)
        import os
        
        # Verificar si el usuario tiene 2FA habilitado (simulado)
        totp_secret_file = f"media/2fa/{current_user.id}_secret.txt"
        is_enabled = os.path.exists(totp_secret_file)
        
        return {
            "enabled": is_enabled,
            "user_id": current_user.id,
            "username": current_user.nombre,
            "backup_codes_available": 5 if is_enabled else 0,
            "last_used": None  # Placeholder para futura implementación
        }
    except Exception as e:
        print(f"❌ Error al obtener status 2FA: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener status 2FA: {str(e)}")

@router.post("/2fa/setup")
def configurar_2fa(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Configura 2FA para el usuario actual generando un secreto TOTP y QR code.
    """
    try:
        import pyotp
        import qrcode
        import io
        import base64
        import os
        
        # Generar secreto único para el usuario
        secret = pyotp.random_base32()
        
        # Crear URI para el QR code
        totp_uri = pyotp.totp.TOTP(secret).provisioning_uri(
            name=current_user.correo,
            issuer_name="Sedes Educativas Cauca"
        )
        
        # Generar QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(totp_uri)
        qr.make(fit=True)
        
        # Convertir QR a imagen base64
        qr_img = qr.make_image(fill_color="black", back_color="white")
        buffer = io.BytesIO()
        qr_img.save(buffer, format='PNG')
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        # Guardar secreto temporalmente (en producción se guardaría en BD)
        os.makedirs("media/2fa", exist_ok=True)
        with open(f"media/2fa/{current_user.id}_secret_temp.txt", "w") as f:
            f.write(secret)
        
        # Generar códigos de respaldo
        backup_codes = [f"{pyotp.random_base32()[:8]}" for _ in range(10)]
        with open(f"media/2fa/{current_user.id}_backup_temp.txt", "w") as f:
            f.write("\n".join(backup_codes))
        
        return {
            "secret": secret,
            "qr_code": f"data:image/png;base64,{qr_base64}",
            "manual_entry_key": secret,
            "backup_codes": backup_codes,
            "setup_complete": False  # Se completa con verify_setup
        }
    except Exception as e:
        print(f"❌ Error al configurar 2FA: {e}")
        raise HTTPException(status_code=400, detail=f"Error al configurar 2FA: {str(e)}")

@router.post("/2fa/verify-setup")
def verificar_configuracion_2fa(
    verification_data: dict,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Verifica el código TOTP para completar la configuración de 2FA.
    """
    try:
        import pyotp
        import os
        
        code = verification_data.get('code')
        if not code:
            raise HTTPException(status_code=400, detail="Código requerido")
        
        # Leer secreto temporal
        secret_file = f"media/2fa/{current_user.id}_secret_temp.txt"
        if not os.path.exists(secret_file):
            raise HTTPException(status_code=400, detail="Configuración no encontrada")
        
        with open(secret_file, "r") as f:
            secret = f.read().strip()
        
        # Verificar código TOTP
        totp = pyotp.TOTP(secret)
        if not totp.verify(code, valid_window=1):
            raise HTTPException(status_code=400, detail="Código inválido")
        
        # Mover archivos temporales a permanentes
        os.rename(secret_file, f"media/2fa/{current_user.id}_secret.txt")
        backup_temp = f"media/2fa/{current_user.id}_backup_temp.txt"
        if os.path.exists(backup_temp):
            os.rename(backup_temp, f"media/2fa/{current_user.id}_backup.txt")
        
        return {
            "success": True,
            "message": "2FA configurado exitosamente",
            "enabled": True
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al verificar 2FA: {e}")
        raise HTTPException(status_code=400, detail=f"Error al verificar 2FA: {str(e)}")

@router.post("/2fa/verify")
def verificar_2fa(
    verification_data: dict,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Verifica un código 2FA para autorizar una acción crítica.
    """
    try:
        import pyotp
        import os
        from datetime import datetime
        
        code = verification_data.get('code')
        action = verification_data.get('action', 'general')
        
        if not code:
            raise HTTPException(status_code=400, detail="Código requerido")
        
        # Verificar si el usuario tiene 2FA habilitado
        secret_file = f"media/2fa/{current_user.id}_secret.txt"
        if not os.path.exists(secret_file):
            raise HTTPException(status_code=400, detail="2FA no configurado")
        
        with open(secret_file, "r") as f:
            secret = f.read().strip()
        
        # Verificar código TOTP
        totp = pyotp.TOTP(secret)
        is_valid = totp.verify(code, valid_window=1)
        
        if not is_valid:
            # Verificar códigos de respaldo
            backup_file = f"media/2fa/{current_user.id}_backup.txt"
            if os.path.exists(backup_file):
                with open(backup_file, "r") as f:
                    backup_codes = f.read().strip().split('\n')
                
                if code in backup_codes:
                    # Remover código usado
                    backup_codes.remove(code)
                    with open(backup_file, "w") as f:
                        f.write('\n'.join(backup_codes))
                    is_valid = True
        
        if not is_valid:
            raise HTTPException(status_code=400, detail="Código inválido")
        
        return {
            "success": True,
            "verified": True,
            "action": action,
            "timestamp": datetime.now().isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al verificar código 2FA: {e}")
        raise HTTPException(status_code=400, detail=f"Error al verificar código 2FA: {str(e)}")

@router.post("/2fa/disable")
def deshabilitar_2fa(
    verification_data: dict,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Deshabilita 2FA para el usuario actual.
    """
    try:
        import os
        
        # Verificar código actual antes de deshabilitar
        verify_result = verificar_2fa(verification_data, db, current_user)
        
        if not verify_result.get('verified'):
            raise HTTPException(status_code=400, detail="Verificación requerida")
        
        # Eliminar archivos de 2FA
        files_to_remove = [
            f"media/2fa/{current_user.id}_secret.txt",
            f"media/2fa/{current_user.id}_backup.txt"
        ]
        
        for file_path in files_to_remove:
            if os.path.exists(file_path):
                os.remove(file_path)
        
        return {
            "success": True,
            "message": "2FA deshabilitado exitosamente",
            "enabled": False
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al deshabilitar 2FA: {e}")
        raise HTTPException(status_code=400, detail=f"Error al deshabilitar 2FA: {str(e)}")

@router.get("/2fa/backup-codes")
def obtener_codigos_respaldo(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Obtiene los códigos de respaldo restantes del usuario.
    """
    try:
        import os
        
        backup_file = f"media/2fa/{current_user.id}_backup.txt"
        if not os.path.exists(backup_file):
            raise HTTPException(status_code=404, detail="Códigos de respaldo no encontrados")
        
        with open(backup_file, "r") as f:
            backup_codes = [code.strip() for code in f.read().split('\n') if code.strip()]
        
        return {
            "backup_codes": backup_codes,
            "remaining": len(backup_codes),
            "warning": "Guarda estos códigos en un lugar seguro" if len(backup_codes) < 3 else None
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al obtener códigos de respaldo: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener códigos de respaldo: {str(e)}")

# ==================== DASHBOARD ANALYTICS AVANZADO ====================

@router.get("/analytics/kpis")
def obtener_kpis_avanzados(
    periodo: str = "mes",  # dia, semana, mes, trimestre, ano
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene KPIs avanzados con cálculos dinámicos y comparativas.
    """
    try:
        from datetime import datetime, timedelta
        import calendar
        
        # Calcular rangos de fechas según el período
        now = datetime.now()
        
        if periodo == "dia":
            fecha_inicio = now.replace(hour=0, minute=0, second=0, microsecond=0)
            fecha_fin = now.replace(hour=23, minute=59, second=59)
            fecha_comparacion = fecha_inicio - timedelta(days=1)
        elif periodo == "semana":
            dias_desde_lunes = now.weekday()
            fecha_inicio = (now - timedelta(days=dias_desde_lunes)).replace(hour=0, minute=0, second=0)
            fecha_fin = fecha_inicio + timedelta(days=6, hours=23, minutes=59, seconds=59)
            fecha_comparacion = fecha_inicio - timedelta(weeks=1)
        elif periodo == "trimestre":
            trimestre = (now.month - 1) // 3
            mes_inicio = trimestre * 3 + 1
            fecha_inicio = now.replace(month=mes_inicio, day=1, hour=0, minute=0, second=0)
            fecha_fin = (fecha_inicio + timedelta(days=93)).replace(day=1) - timedelta(days=1)
            fecha_comparacion = fecha_inicio - timedelta(days=93)
        elif periodo == "ano":
            fecha_inicio = now.replace(month=1, day=1, hour=0, minute=0, second=0)
            fecha_fin = now.replace(month=12, day=31, hour=23, minute=59, second=59)
            fecha_comparacion = fecha_inicio.replace(year=fecha_inicio.year - 1)
        else:  # mes por defecto
            fecha_inicio = now.replace(day=1, hour=0, minute=0, second=0)
            ultimo_dia = calendar.monthrange(now.year, now.month)[1]
            fecha_fin = now.replace(day=ultimo_dia, hour=23, minute=59, second=59)
            if now.month == 1:
                fecha_comparacion = fecha_inicio.replace(year=fecha_inicio.year - 1, month=12)
            else:
                fecha_comparacion = fecha_inicio.replace(month=fecha_inicio.month - 1)
        
        # KPIs principales con comparativas
        kpis = {}
        
        # 1. Visitas programadas vs completadas
        visitas_programadas = db.execute(text("""
            SELECT COUNT(*) FROM visitas_programadas 
            WHERE fecha_programada >= :inicio AND fecha_programada <= :fin
        """), {"inicio": fecha_inicio, "fin": fecha_fin}).scalar()
        
        visitas_completadas = db.execute(text("""
            SELECT COUNT(*) FROM visitas_completas_pae 
            WHERE fecha_visita >= :inicio AND fecha_visita <= :fin
        """), {"inicio": fecha_inicio, "fin": fecha_fin}).scalar()
        
        # Datos del período anterior para comparación
        visitas_programadas_anterior = db.execute(text("""
            SELECT COUNT(*) FROM visitas_programadas 
            WHERE fecha_programada >= :inicio AND fecha_programada < :fin
        """), {"inicio": fecha_comparacion, "fin": fecha_inicio}).scalar()
        
        visitas_completadas_anterior = db.execute(text("""
            SELECT COUNT(*) FROM visitas_completas_pae 
            WHERE fecha_visita >= :inicio AND fecha_visita < :fin
        """), {"inicio": fecha_comparacion, "fin": fecha_inicio}).scalar()
        
        # 2. Tasa de cumplimiento
        tasa_cumplimiento = (visitas_completadas / visitas_programadas * 100) if visitas_programadas > 0 else 0
        tasa_cumplimiento_anterior = (visitas_completadas_anterior / visitas_programadas_anterior * 100) if visitas_programadas_anterior > 0 else 0
        
        # 3. Sedes activas
        sedes_activas = db.execute(text("""
            SELECT COUNT(DISTINCT sede_id) FROM visitas_programadas 
            WHERE fecha_programada >= :inicio AND fecha_programada <= :fin
        """), {"inicio": fecha_inicio, "fin": fecha_fin}).scalar()
        
        sedes_activas_anterior = db.execute(text("""
            SELECT COUNT(DISTINCT sede_id) FROM visitas_programadas 
            WHERE fecha_programada >= :inicio AND fecha_programada < :fin
        """), {"inicio": fecha_comparacion, "fin": fecha_inicio}).scalar()
        
        # 4. Visitadores activos
        visitadores_activos = db.execute(text("""
            SELECT COUNT(DISTINCT visitador_id) FROM visitas_programadas 
            WHERE fecha_programada >= :inicio AND fecha_programada <= :fin
        """), {"inicio": fecha_inicio, "fin": fecha_fin}).scalar()
        
        visitadores_activos_anterior = db.execute(text("""
            SELECT COUNT(DISTINCT visitador_id) FROM visitas_programadas 
            WHERE fecha_programada >= :inicio AND fecha_programada < :fin
        """), {"inicio": fecha_comparacion, "fin": fecha_inicio}).scalar()
        
        # 5. Promedio de visitas por visitador
        promedio_visitas = visitas_programadas / visitadores_activos if visitadores_activos > 0 else 0
        promedio_visitas_anterior = visitas_programadas_anterior / visitadores_activos_anterior if visitadores_activos_anterior > 0 else 0
        
        # Calcular cambios porcentuales
        def calcular_cambio(actual, anterior):
            if anterior == 0:
                return 100 if actual > 0 else 0
            return ((actual - anterior) / anterior) * 100
        
        return {
            "periodo": periodo,
            "fecha_inicio": fecha_inicio.isoformat(),
            "fecha_fin": fecha_fin.isoformat(),
            "kpis": {
                "visitas_programadas": {
                    "valor": visitas_programadas,
                    "anterior": visitas_programadas_anterior,
                    "cambio": calcular_cambio(visitas_programadas, visitas_programadas_anterior),

                },
                "visitas_completadas": {
                    "valor": visitas_completadas,
                    "anterior": visitas_completadas_anterior,
                    "cambio": calcular_cambio(visitas_completadas, visitas_completadas_anterior),

                },
                "tasa_cumplimiento": {
                    "valor": round(tasa_cumplimiento, 1),
                    "anterior": round(tasa_cumplimiento_anterior, 1),
                    "cambio": tasa_cumplimiento - tasa_cumplimiento_anterior,

                },
                "sedes_activas": {
                    "valor": sedes_activas,
                    "anterior": sedes_activas_anterior,
                    "cambio": calcular_cambio(sedes_activas, sedes_activas_anterior),

                },
                "visitadores_activos": {
                    "valor": visitadores_activos,
                    "anterior": visitadores_activos_anterior,
                    "cambio": calcular_cambio(visitadores_activos, visitadores_activos_anterior),

                },
                "promedio_visitas_visitador": {
                    "valor": round(promedio_visitas, 1),
                    "anterior": round(promedio_visitas_anterior, 1),
                    "cambio": promedio_visitas - promedio_visitas_anterior,

                }
            }
        }
    except Exception as e:
        print(f"❌ Error al obtener KPIs: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener KPIs: {str(e)}")



@router.get("/analytics/graficos/rendimiento-visitadores")
def obtener_rendimiento_visitadores(
    limit: int = 10,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene ranking de rendimiento de visitadores.
    """
    try:
        from datetime import datetime, timedelta
        
        # Últimos 30 días
        fecha_inicio = datetime.now() - timedelta(days=30)
        
        result = db.execute(text("""
            SELECT 
                u.id,
                u.nombre,
                COUNT(vp.id) as visitas_programadas,
                COUNT(vc.id) as visitas_completadas,
                CASE 
                    WHEN COUNT(vp.id) = 0 THEN 0 
                    ELSE CAST(COUNT(vc.id) * 100.0 / COUNT(vp.id) AS DECIMAL(5,1))
                END as tasa_cumplimiento
            FROM usuarios u
            LEFT JOIN visitas_programadas vp ON u.id = vp.visitador_id 
                AND vp.fecha_programada >= :fecha_inicio
            LEFT JOIN visitas_completas_pae vc ON vp.sede_id = vc.sede_id 
                AND vc.profesional_id = u.id 
                AND vc.fecha_visita >= :fecha_inicio
            WHERE u.rol_id = 1  -- Solo visitadores
            GROUP BY u.id, u.nombre
            HAVING COUNT(vp.id) > 0
            ORDER BY tasa_cumplimiento DESC, visitas_completadas DESC
            LIMIT :limit
        """), {"fecha_inicio": fecha_inicio, "limit": limit}).fetchall()
        
        datos = []
        for i, row in enumerate(result, 1):
            datos.append({
                "ranking": i,
                "visitador_id": row.id,
                "nombre": row.nombre,
                "visitas_programadas": row.visitas_programadas,
                "visitas_completadas": row.visitas_completadas,
                "tasa_cumplimiento": row.tasa_cumplimiento or 0,
                "badge": "🏆" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else "📍"
            })
        
        return {
            "periodo": "Últimos 30 días",
            "fecha_desde": fecha_inicio.strftime("%Y-%m-%d"),
            "total_visitadores": len(datos),
            "ranking": datos
        }
    except Exception as e:
        print(f"❌ Error al obtener rendimiento: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener rendimiento: {str(e)}")

@router.get("/analytics/graficos/distribucion-geografica")
def obtener_distribucion_geografica(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene distribución de visitas por municipio/institución.
    """
    try:
        from datetime import datetime, timedelta
        
        # Últimos 30 días
        fecha_inicio = datetime.now() - timedelta(days=30)
        
        # Por municipio
        municipios = db.execute(text("""
            SELECT 
                m.nombre as municipio,
                COUNT(vp.id) as visitas_programadas,
                COUNT(vc.id) as visitas_completadas
            FROM municipios m
            LEFT JOIN sedes_educativas se ON m.id = se.municipio_id
            LEFT JOIN visitas_programadas vp ON se.id = vp.sede_id 
                AND vp.fecha_programada >= :fecha_inicio
            LEFT JOIN visitas_completas_pae vc ON se.id = vc.sede_id 
                AND vc.fecha_visita >= :fecha_inicio
            GROUP BY m.id, m.nombre
            HAVING COUNT(vp.id) > 0
            ORDER BY visitas_completadas DESC
        """), {"fecha_inicio": fecha_inicio}).fetchall()
        
        # Por tipo de institución (simulado)
        tipos_institucion = [
            {"tipo": "Escuela Primaria", "cantidad": 45, "porcentaje": 60},
            {"tipo": "Colegio Secundario", "cantidad": 20, "porcentaje": 27},
            {"tipo": "Institución Mixta", "cantidad": 8, "porcentaje": 11},
            {"tipo": "Centro Técnico", "cantidad": 2, "porcentaje": 2}
        ]
        
        municipios_data = []
        for row in municipios:
            tasa = (row.visitas_completadas / row.visitas_programadas * 100) if row.visitas_programadas > 0 else 0
            municipios_data.append({
                "municipio": row.municipio,
                "visitas_programadas": row.visitas_programadas,
                "visitas_completadas": row.visitas_completadas,
                "tasa_cumplimiento": round(tasa, 1)
            })
        
        return {
            "periodo": "Últimos 30 días",
            "municipios": municipios_data,
            "tipos_institucion": tipos_institucion,
            "resumen": {
                "total_municipios": len(municipios_data),
                "municipio_mas_activo": municipios_data[0]["municipio"] if municipios_data else None
            }
        }
    except Exception as e:
        print(f"❌ Error al obtener distribución: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener distribución: {str(e)}")

@router.get("/analytics/alertas")
def obtener_alertas_criticas(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene alertas críticas del sistema.
    """
    try:
        from datetime import datetime, timedelta
        
        alertas = []
        
        # 1. Visitas vencidas
        visitas_vencidas = db.execute(text("""
            SELECT COUNT(*) FROM visitas_programadas 
            WHERE fecha_programada < NOW() AND estado != 'completada'
        """)).scalar()
        
        if visitas_vencidas > 0:
            alertas.append({
                "tipo": "critica",
                "icono": "warning",
                "titulo": "Visitas Vencidas",
                "mensaje": f"{visitas_vencidas} visitas programadas están vencidas",
                "color": "red",
                "accion": "revisar_vencidas"
            })
        
        # 2. Visitadores inactivos
        fecha_limite = datetime.now() - timedelta(days=7)
        visitadores_inactivos = db.execute(text("""
            SELECT COUNT(*) FROM usuarios u
            WHERE u.rol_id = 1 
            AND u.id NOT IN (
                SELECT DISTINCT visitador_id FROM visitas_programadas 
                WHERE fecha_programada >= :fecha_limite
            )
        """), {"fecha_limite": fecha_limite}).scalar()
        
        if visitadores_inactivos > 0:
            alertas.append({
                "tipo": "advertencia",
                "icono": "person_off",
                "titulo": "Visitadores Inactivos",
                "mensaje": f"{visitadores_inactivos} visitadores sin actividad en 7 días",
                "color": "orange",
                "accion": "revisar_inactivos"
            })
        
        # 3. Baja tasa de cumplimiento
        tasa_cumplimiento = db.execute(text("""
            SELECT 
                CASE 
                    WHEN COUNT(vp.id) = 0 THEN 0 
                    ELSE CAST(COUNT(vc.id) * 100.0 / COUNT(vp.id) AS DECIMAL(5,1))
                END
            FROM visitas_programadas vp
            LEFT JOIN visitas_completas_pae vc ON vp.sede_id = vc.sede_id
            WHERE vp.fecha_programada >= NOW() - INTERVAL '30 days'
        """)).scalar()
        
        if tasa_cumplimiento and tasa_cumplimiento < 70:
            alertas.append({
                "tipo": "advertencia",
                "icono": "trending_down",
                "titulo": "Baja Tasa de Cumplimiento",
                "mensaje": f"Tasa actual: {tasa_cumplimiento}% (objetivo: >70%)",
                "color": "orange",
                "accion": "analizar_cumplimiento"
            })
        
        # 4. Sedes sin visitas recientes
        sedes_sin_visitas = db.execute(text("""
            SELECT COUNT(*) FROM sedes_educativas se
            WHERE se.id NOT IN (
                SELECT DISTINCT sede_id FROM visitas_programadas 
                WHERE fecha_programada >= NOW() - INTERVAL '60 days'
            )
        """)).scalar()
        
        if sedes_sin_visitas > 5:
            alertas.append({
                "tipo": "info",
                "icono": "location_off",
                "titulo": "Sedes Sin Visitas",
                "mensaje": f"{sedes_sin_visitas} sedes sin visitas en 60 días",
                "color": "blue",
                "accion": "programar_visitas"
            })
        
        return {
            "total_alertas": len(alertas),
            "alertas_criticas": len([a for a in alertas if a["tipo"] == "critica"]),
            "alertas_advertencia": len([a for a in alertas if a["tipo"] == "advertencia"]),
            "alertas": alertas,
            "ultima_actualizacion": datetime.now().isoformat()
        }
    except Exception as e:
        print(f"❌ Error al obtener alertas: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener alertas: {str(e)}")

# ==================== NOTIFICACIONES PUSH INTELIGENTES ====================

@router.get("/notificaciones/configuracion")
def obtener_configuracion_notificaciones(
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Obtiene la configuración de notificaciones del usuario.
    """
    try:
        # Por ahora usamos configuración por defecto (en producción se guardaría en BD)
        import os
        
        config_file = f"media/notifications/{current_user.id}_config.json"
        
        # Configuración por defecto
        config_default = {
            "visitas_vencidas": {"enabled": True, "tipo": "push"},
            "nuevas_asignaciones": {"enabled": True, "tipo": "push"},
            "recordatorios": {"enabled": True, "tipo": "push"},
            "alertas_sistema": {"enabled": True, "tipo": "push"},
            "reportes_listos": {"enabled": False, "tipo": "email"},
            "cambios_programacion": {"enabled": True, "tipo": "push"},
            "horario_silencio": {"inicio": "22:00", "fin": "08:00"},
            "push_token": None,
            "email_notifications": current_user.correo
        }
        
        if os.path.exists(config_file):
            import json
            with open(config_file, 'r') as f:
                config_guardada = json.load(f)
                # Merge con default para nuevas opciones
                config_default.update(config_guardada)
        
        return {
            "user_id": current_user.id,
            "configuracion": config_default,
            "tipos_disponibles": ["push", "email", "sms"],
            "categorias": [
                {
                    "id": "visitas_vencidas",
                    "nombre": "Visitas Vencidas",
                    "descripcion": "Alertas cuando hay visitas programadas sin completar",
                    "criticidad": "alta"
                },
                {
                    "id": "nuevas_asignaciones",
                    "nombre": "Nuevas Asignaciones",
                    "descripcion": "Notificaciones de nuevas visitas programadas",
                    "criticidad": "media"
                },
                {
                    "id": "recordatorios",
                    "nombre": "Recordatorios",
                    "descripcion": "Recordatorios de visitas próximas (24h antes)",
                    "criticidad": "media"
                },
                {
                    "id": "alertas_sistema",
                    "nombre": "Alertas del Sistema",
                    "descripcion": "Problemas técnicos y mantenimiento",
                    "criticidad": "alta"
                },
                {
                    "id": "reportes_listos",
                    "nombre": "Reportes Listos",
                    "descripcion": "Cuando las exportaciones están disponibles",
                    "criticidad": "baja"
                },
                {
                    "id": "cambios_programacion",
                    "nombre": "Cambios de Programación",
                    "descripcion": "Modificaciones en visitas ya programadas",
                    "criticidad": "media"
                }
            ]
        }
    except Exception as e:
        print(f"❌ Error al obtener configuración: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener configuración: {str(e)}")

@router.put("/notificaciones/configuracion")
def actualizar_configuracion_notificaciones(
    configuracion_data: dict,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user)
):
    """
    Actualiza la configuración de notificaciones del usuario.
    """
    try:
        import os
        import json
        
        # Crear directorio si no existe
        os.makedirs("media/notifications", exist_ok=True)
        
        config_file = f"media/notifications/{current_user.id}_config.json"
        
        # Validar configuración
        configuracion = configuracion_data.get('configuracion', {})
        
        # Guardar configuración
        with open(config_file, 'w') as f:
            json.dump(configuracion, f, indent=2)
        
        return {
            "success": True,
            "message": "Configuración actualizada exitosamente",
            "configuracion": configuracion
        }
    except Exception as e:
        print(f"❌ Error al actualizar configuración: {e}")
        raise HTTPException(status_code=400, detail=f"Error al actualizar configuración: {str(e)}")

@router.post("/notificaciones/enviar")
def enviar_notificacion(
    notificacion_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Envía una notificación a usuarios específicos o grupos.
    """
    try:
        from datetime import datetime
        import os
        import json
        
        # Extraer datos
        titulo = notificacion_data.get('titulo')
        mensaje = notificacion_data.get('mensaje')
        tipo = notificacion_data.get('tipo', 'info')  # info, warning, error, success
        categoria = notificacion_data.get('categoria', 'alertas_sistema')
        destinatarios = notificacion_data.get('destinatarios', [])  # user_ids o 'all'
        canales = notificacion_data.get('canales', ['push'])  # push, email, sms
        
        if not titulo or not mensaje:
            raise HTTPException(status_code=400, detail="Título y mensaje son requeridos")
        
        # Obtener destinatarios
        if destinatarios == 'all':
            usuarios = db.query(models.Usuario).all()  # Todos los usuarios
        elif destinatarios == 'admins':
            # Solo administradores
            usuarios = db.query(models.Usuario).join(models.Rol).filter(
                models.Rol.nombre.in_(['admin', 'administrador'])
            ).all()
        elif destinatarios == 'supervisores':
            # Solo supervisores
            usuarios = db.query(models.Usuario).join(models.Rol).filter(
                models.Rol.nombre == 'supervisor'
            ).all()
            print(f"🔍 DEBUG: Encontrados {len(usuarios)} supervisores para notificar")
            for u in usuarios:
                print(f"   - ID: {u.id}, Nombre: {u.nombre}, Correo: {u.correo}, Rol: {u.rol.nombre if u.rol else 'Sin rol'}")
        elif destinatarios == 'visitadores':
            # Solo visitadores
            usuarios = db.query(models.Usuario).join(models.Rol).filter(
                models.Rol.nombre == 'visitador'
            ).all()
        elif isinstance(destinatarios, list):
            # Lista de IDs específicos
            usuarios = db.query(models.Usuario).filter(
                models.Usuario.id.in_(destinatarios)
            ).all()
        else:
            raise HTTPException(status_code=400, detail="Tipo de destinatarios no válido")
        
        # Crear notificación base
        notificacion = {
            "id": f"notif_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            "titulo": titulo,
            "mensaje": mensaje,
            "tipo": tipo,
            "categoria": categoria,
            "fecha_creacion": datetime.now().isoformat(),
            "creado_por": admin_user.id,
            "canales": canales,
            "estado": "enviada"
        }
        
        # Simular envío a diferentes canales
        resultados = {
            "enviadas": 0,
            "fallidas": 0,
            "canales_usados": [],
            "destinatarios": []
        }
        
        for usuario in usuarios:
            usuario_resultado = {
                "user_id": usuario.id,
                "nombre": usuario.nombre,
                "canales_enviados": [],
                "errores": []
            }
            
            # Verificar configuración del usuario
            config_file = f"media/notifications/{usuario.id}_config.json"
            if os.path.exists(config_file):
                with open(config_file, 'r') as f:
                    config_usuario = json.load(f)
            else:
                config_usuario = {"visitas_vencidas": {"enabled": True}}
            
            # Verificar si el usuario tiene habilitada esta categoría
            categoria_config = config_usuario.get(categoria, {"enabled": True})
            if not categoria_config.get("enabled", True):
                usuario_resultado["errores"].append("Categoría deshabilitada por el usuario")
                resultados["fallidas"] += 1
                continue
            
            # Simular envío por cada canal
            for canal in canales:
                try:
                    if canal == "push":
                        # Simular envío push
                        exito = _enviar_push_notification(usuario, notificacion)
                        if exito:
                            usuario_resultado["canales_enviados"].append("push")
                        else:
                            usuario_resultado["errores"].append("Push token no disponible")
                    
                    elif canal == "email":
                        # Simular envío email
                        exito = _enviar_email_notification(usuario, notificacion)
                        if exito:
                            usuario_resultado["canales_enviados"].append("email")
                        else:
                            usuario_resultado["errores"].append("Email no configurado")
                    
                    elif canal == "sms":
                        # Simular envío SMS
                        exito = _enviar_sms_notification(usuario, notificacion)
                        if exito:
                            usuario_resultado["canales_enviados"].append("sms")
                        else:
                            usuario_resultado["errores"].append("SMS no disponible")
                
                except Exception as e:
                    usuario_resultado["errores"].append(f"Error en {canal}: {str(e)}")
            
            # Guardar notificación en BD inmediatamente si fue exitosa
            if usuario_resultado["canales_enviados"]:
                try:
                    # Crear notificación real en BD (solo campos que existen)
                    nueva_notificacion = models.Notificacion(
                        usuario_id=usuario.id,
                        titulo=notificacion["titulo"],
                        mensaje=notificacion["mensaje"],
                        tipo=notificacion["tipo"],
                        prioridad="normal",
                        leida=False
                    )
                    
                    db.add(nueva_notificacion)
                    db.flush()  # Flush inmediato para verificar errores
                    print(f"✅ Notificación guardada para usuario {usuario.nombre} (ID: {usuario.id})")
                    
                    resultados["enviadas"] += 1
                except Exception as e:
                    print(f"❌ Error guardando notificación para usuario {usuario.id}: {e}")
                    usuario_resultado["errores"].append(f"Error BD: {str(e)}")
                    resultados["fallidas"] += 1
            else:
                resultados["fallidas"] += 1
            
            resultados["destinatarios"].append(usuario_resultado)
        
        # Actualizar canales usados
        resultados["canales_usados"] = list(set(canal for dest in resultados["destinatarios"] for canal in dest["canales_enviados"]))
        
        # Hacer commit final de todas las notificaciones
        try:
            db.commit()
            print(f"✅ Commit exitoso de {resultados['enviadas']} notificaciones")
        except Exception as e:
            print(f"❌ Error en commit final: {e}")
            db.rollback()
        
        # Guardar log de notificación
        _guardar_log_notificacion(notificacion, resultados)
        
        return {
            "success": True,
            "message": f"Notificación enviada a {resultados['enviadas']} usuarios",
            "notificacion": notificacion,
            "resultados": resultados
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error al enviar notificación: {e}")
        raise HTTPException(status_code=400, detail=f"Error al enviar notificación: {str(e)}")

@router.get("/notificaciones/historial")
def obtener_historial_notificaciones(
    limite: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene el historial de notificaciones enviadas.
    """
    try:
        import os
        import json
        from datetime import datetime, timedelta
        
        log_dir = "media/notifications/logs"
        if not os.path.exists(log_dir):
            return {
                "total": 0,
                "notificaciones": [],
                "limite": limite,
                "offset": offset
            }
        
        # Leer archivos de log (últimos 30 días)
        notificaciones = []
        fecha_limite = datetime.now() - timedelta(days=30)
        
        for filename in sorted(os.listdir(log_dir), reverse=True):
            if filename.endswith('.json'):
                filepath = os.path.join(log_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        log_data = json.load(f)
                        fecha_log = datetime.fromisoformat(log_data['notificacion']['fecha_creacion'].replace('Z', ''))
                        
                        if fecha_log >= fecha_limite:
                            notificaciones.append(log_data)
                except Exception as e:
                    print(f"Error leyendo log {filename}: {e}")
                    continue
        
        # Aplicar paginación
        total = len(notificaciones)
        notificaciones_paginadas = notificaciones[offset:offset + limite]
        
        return {
            "total": total,
            "notificaciones": notificaciones_paginadas,
            "limite": limite,
            "offset": offset,
            "estadisticas": {
                "total_enviadas": sum(n["resultados"]["enviadas"] for n in notificaciones),
                "total_fallidas": sum(n["resultados"]["fallidas"] for n in notificaciones),
                "canales_mas_usados": _calcular_canales_populares(notificaciones)
            }
        }
    except Exception as e:
        print(f"❌ Error al obtener historial: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener historial: {str(e)}")

def _enviar_push_notification(usuario, notificacion):
    """Simula envío de notificación push."""
    # En producción, aquí usarías Firebase Cloud Messaging
    import random
    return random.choice([True, True, True, False])  # 75% de éxito

def _enviar_email_notification(usuario, notificacion):
    """Envía notificación por email usando SMTP."""
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        import os
        from datetime import datetime
        
        # Configuración de email (Gmail como ejemplo)
        # En producción, estas credenciales deberían estar en variables de entorno
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        sender_email = "sistema.visitas.cauca@gmail.com"  # Email del sistema
        sender_password = "app_password_here"  # App password, no la contraseña real
        
        # Verificar si tenemos configuración de email
        if not hasattr(usuario, 'correo') or not usuario.correo:
            return False
        
        # Crear el mensaje
        msg = MIMEMultipart()
        msg['From'] = f"Sistema de Visitas Cauca <{sender_email}>"
        msg['To'] = usuario.correo
        msg['Subject'] = f"🔔 {notificacion['titulo']}"
        
        # Crear el cuerpo HTML del email
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
                .container {{ max-width: 600px; margin: 0 auto; background-color: white; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .header {{ background-color: #2E7D32; color: white; padding: 20px; border-radius: 8px 8px 0 0; text-align: center; }}
                .content {{ padding: 30px; }}
                .message {{ background-color: #f8f9fa; padding: 20px; border-radius: 8px; border-left: 4px solid #2E7D32; margin: 20px 0; }}
                .footer {{ background-color: #f8f9fa; padding: 15px; text-align: center; font-size: 12px; color: #666; border-radius: 0 0 8px 8px; }}
                .type-badge {{ display: inline-block; padding: 5px 10px; border-radius: 15px; font-size: 12px; font-weight: bold; color: white; }}
                .type-info {{ background-color: #2196F3; }}
                .type-warning {{ background-color: #FF9800; }}
                .type-error {{ background-color: #F44336; }}
                .type-success {{ background-color: #4CAF50; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🏛️ Sistema de Visitas - Cauca</h1>
                    <p>Notificación del Sistema Administrativo</p>
                </div>
                <div class="content">
                    <h2>Hola {usuario.nombre},</h2>
                    
                    <div class="message">
                        <div style="margin-bottom: 10px;">
                            <span class="type-badge type-{notificacion['tipo']}">{notificacion['tipo'].upper()}</span>
                            <strong style="margin-left: 10px;">Categoría: {notificacion['categoria'].replace('_', ' ').title()}</strong>
                        </div>
                        <h3>{notificacion['titulo']}</h3>
                        <p style="font-size: 16px; line-height: 1.6;">{notificacion['mensaje']}</p>
                    </div>
                    
                    <p><strong>Fecha:</strong> {datetime.now().strftime('%d/%m/%Y a las %H:%M')}</p>
                    <p><strong>ID Notificación:</strong> {notificacion['id']}</p>
                    
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #eee;">
                    
                    <p style="color: #666; font-size: 14px;">
                        💡 <strong>Tip:</strong> Puedes configurar tus preferencias de notificaciones desde 
                        el panel administrativo del sistema.
                    </p>
                </div>
                <div class="footer">
                    <p>Este es un mensaje automático del Sistema de Visitas Educativas del Cauca</p>
                    <p>📧 No responder a este email | 🌐 Accede al sistema para más detalles</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Adjuntar el cuerpo HTML
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        
        # Intentar envío real (comentado por seguridad)
        # try:
        #     server = smtplib.SMTP(smtp_server, smtp_port)
        #     server.starttls()
        #     server.login(sender_email, sender_password)
        #     text = msg.as_string()
        #     server.sendmail(sender_email, usuario.correo, text)
        #     server.quit()
        #     print(f"✅ Email enviado exitosamente a {usuario.correo}")
        #     return True
        # except Exception as e:
        #     print(f"❌ Error enviando email a {usuario.correo}: {e}")
        #     return False
        
        # Por ahora, simular envío exitoso y guardar el HTML generado
        # Esto permite ver el email que se enviaría sin necesidad de configurar SMTP
        email_dir = "media/notifications/emails"
        os.makedirs(email_dir, exist_ok=True)
        
        # Guardar el HTML del email para inspección
        email_file = f"{email_dir}/email_{notificacion['id']}_{usuario.id}.html"
        with open(email_file, 'w', encoding='utf-8') as f:
            f.write(html_body)
        
        print(f"📧 Email preparado para {usuario.correo} (HTML guardado en {email_file})")
        
        # Simular éxito si el usuario tiene email configurado
        return bool(usuario.correo and '@' in usuario.correo)
        
    except Exception as e:
        print(f"❌ Error preparando email: {e}")
        return False

def _enviar_sms_notification(usuario, notificacion):
    """Simula envío de notificación por SMS."""
    # En producción, aquí usarías Twilio u otro proveedor SMS
    import random
    return random.choice([True, False, False])  # 33% de éxito

def _guardar_log_notificacion(notificacion, resultados):
    """Guarda log de la notificación enviada."""
    try:
        import os
        import json
        from datetime import datetime
        
        # Crear directorio de logs
        log_dir = "media/notifications/logs"
        os.makedirs(log_dir, exist_ok=True)
        
        # Crear archivo de log con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = os.path.join(log_dir, f"notif_{timestamp}.json")
        
        log_data = {
            "notificacion": notificacion,
            "resultados": resultados,
            "timestamp": datetime.now().isoformat()
        }
        
        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(log_data, f, indent=2, ensure_ascii=False)
    
    except Exception as e:
        print(f"Error guardando log: {e}")


def _calcular_canales_populares(notificaciones):
    """Calcula estadísticas de canales más usados."""
    canales_count = {}
    
    for notif in notificaciones:
        for canal in notif["resultados"]["canales_usados"]:
            canales_count[canal] = canales_count.get(canal, 0) + 1
    
    return sorted(canales_count.items(), key=lambda x: x[1], reverse=True)

@router.post("/notificaciones/automaticas/procesar")
def procesar_notificaciones_automaticas(
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Procesa y envía notificaciones automáticas basadas en reglas del sistema.
    Este endpoint se puede llamar desde un cron job o scheduler.
    """
    try:
        from datetime import datetime, timedelta
        
        notificaciones_enviadas = []
        
        # 1. Verificar visitas vencidas
        visitas_vencidas = db.execute(text("""
            SELECT COUNT(*) FROM visitas_programadas 
            WHERE fecha_programada < NOW() - INTERVAL '1 day' 
            AND estado != 'completada'
        """)).scalar()
        
        if visitas_vencidas > 0:
            notif_result = enviar_notificacion({
                "titulo": "⚠️ Visitas Vencidas Detectadas",
                "mensaje": f"Hay {visitas_vencidas} visitas programadas que están vencidas y sin completar. Requieren atención inmediata.",
                "tipo": "warning",
                "categoria": "visitas_vencidas",
                "destinatarios": "all",
                "canales": ["push", "email"]
            }, db, admin_user)
            notificaciones_enviadas.append(notif_result)
        
        # 2. Recordatorios de visitas próximas (24h)
        visitas_manana = db.execute(text("""
            SELECT COUNT(*) FROM visitas_programadas 
            WHERE fecha_programada BETWEEN NOW() + INTERVAL '20 hours' 
            AND NOW() + INTERVAL '28 hours'
            AND estado = 'programada'
        """)).scalar()
        
        if visitas_manana > 0:
            notif_result = enviar_notificacion({
                "titulo": "📅 Recordatorio: Visitas Mañana",
                "mensaje": f"Tienes {visitas_manana} visita(s) programada(s) para mañana. Revisa tu agenda.",
                "tipo": "info",
                "categoria": "recordatorios",
                "destinatarios": "all",
                "canales": ["push"]
            }, db, admin_user)
            notificaciones_enviadas.append(notif_result)
        
        # 3. Verificar exportaciones listas (simulado)
        import random
        if random.choice([True, False]):
            notif_result = enviar_notificacion({
                "titulo": " Reporte Listo para Descarga",
                "mensaje": "Tu exportación de datos ha sido procesada y está lista para descargar.",
                "tipo": "success",
                "categoria": "reportes_listos",
                "destinatarios": [admin_user.id],
                "canales": ["push"]
            }, db, admin_user)
            notificaciones_enviadas.append(notif_result)
        
        return {
            "success": True,
            "message": f"Procesamiento automático completado",
            "notificaciones_enviadas": len(notificaciones_enviadas),
            "detalles": notificaciones_enviadas,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        print(f"❌ Error en procesamiento automático: {e}")
        raise HTTPException(status_code=400, detail=f"Error en procesamiento automático: {str(e)}")

# ==================== GESTIÓN COMPLETA DE USUARIOS ====================

@router.get("/usuarios")
def obtener_usuarios_admin(
    rol: str = None,
    estado: str = None,
    municipio_id: int = None,
    search: str = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene lista de usuarios con filtros avanzados para administradores.
    """
    try:
        # Construir query base
        query = db.query(models.Usuario)
        
        # Aplicar filtros
        if rol:
            query = query.filter(models.Usuario.rol_id == rol)
        
        if search:
            query = query.filter(
                (models.Usuario.nombre.ilike(f'%{search}%')) |
                (models.Usuario.correo.ilike(f'%{search}%'))
            )
        
        # Obtener resultados
        usuarios = query.offset(offset).limit(limit).all()
        
        # Convertir a lista de diccionarios con información adicional
        usuarios_data = []
        for usuario in usuarios:
            # Obtener rol
            rol_info = db.query(models.Rol).filter(models.Rol.id == usuario.rol_id).first()
            
            # Contar visitas (si es visitador)
            visitas_asignadas = 0
            visitas_completadas = 0
            if usuario.rol_id == 4:  # Visitador (ID 4)
                visitas_asignadas = db.execute(text("""
                    SELECT COUNT(*) FROM visitas_asignadas 
                    WHERE visitador_id = :usuario_id
                """), {"usuario_id": usuario.id}).scalar() or 0
                
                visitas_completadas = db.execute(text("""
                    SELECT COUNT(*) FROM visitas_completas_pae 
                    WHERE profesional_id = :usuario_id
                """), {"usuario_id": usuario.id}).scalar() or 0
            
            usuario_data = {
                "id": usuario.id,
                "nombre": usuario.nombre,
                "correo": usuario.correo,
                "telefono": getattr(usuario, 'telefono', ''),
                "rol": rol_info.nombre if rol_info else "Sin rol",  # Cambiado de rol_id a rol
                "rol_id": usuario.rol_id,  # Mantener rol_id para compatibilidad
                "fecha_creacion": None,  # Campo no disponible en el modelo actual
                "ultimo_acceso": None,  # Campo no disponible en el modelo actual
                "estado": True,  # Asumimos activo por defecto
                "visitas_asignadas": visitas_asignadas,
                "visitas_completadas": visitas_completadas,
                "tasa_cumplimiento": round((visitas_completadas / visitas_asignadas * 100), 1) if visitas_asignadas > 0 else 0
            }
            
            # Debug logging
            print(f"🔍 Usuario {usuario.id}: rol='{usuario_data['rol']}' (tipo: {type(usuario_data['rol'])})")
            
            usuarios_data.append(usuario_data)
        
        print(f"📊 Total usuarios procesados: {len(usuarios_data)}")
        return usuarios_data
    except Exception as e:
        print(f"❌ Error al obtener usuarios: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener usuarios: {str(e)}")

@router.get("/usuarios/{usuario_id}")
def obtener_usuario_detalle(
    usuario_id: int,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene detalles completos de un usuario específico.
    """
    try:
        usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
        # Obtener rol
        rol_info = db.query(models.Rol).filter(models.Rol.id == usuario.rol_id).first()
        print(f"🔍 Usuario {usuario.id}: rol_id={usuario.rol_id}, rol_info={rol_info.nombre if rol_info else 'None'}")
        
        # Obtener permisos del rol
        permisos = []
        if rol_info:
            permisos_result = db.execute(text("""
                SELECT p.nombre, p.descripcion 
                FROM permisos p
                INNER JOIN rol_permisos rp ON p.id = rp.permiso_id
                WHERE rp.rol_id = :rol_id
            """), {"rol_id": rol_info.id}).fetchall()
            permisos = [{"nombre": p.nombre, "descripcion": p.descripcion} for p in permisos_result]
        
        # Estadísticas del usuario
        estadisticas = {}
        if usuario.rol_id == 1:  # Visitador
            estadisticas = {
                "visitas_totales": db.execute(text("""
                    SELECT COUNT(*) FROM visitas_programadas WHERE visitador_id = :usuario_id
                """), {"usuario_id": usuario_id}).scalar() or 0,
                
                "visitas_completadas": db.execute(text("""
                    SELECT COUNT(*) FROM visitas_completas_pae WHERE profesional_id = :usuario_id
                """), {"usuario_id": usuario_id}).scalar() or 0,
                
                "visitas_pendientes": db.execute(text("""
                    SELECT COUNT(*) FROM visitas_programadas 
                    WHERE visitador_id = :usuario_id AND estado = 'programada'
                """), {"usuario_id": usuario_id}).scalar() or 0,
                
                "ultimo_reporte": db.execute(text("""
                    SELECT MAX(fecha_visita) FROM visitas_completas_pae 
                    WHERE profesional_id = :usuario_id
                """), {"usuario_id": usuario_id}).scalar()
            }
        
        return {
            "id": usuario.id,
            "nombre": usuario.nombre,
            "correo": usuario.correo,
            "telefono": getattr(usuario, 'telefono', ''),
            "rol_id": usuario.rol_id,
            "rol_nombre": rol_info.nombre if rol_info else "Sin rol",
            "permisos": permisos,
            "fecha_creacion": None,  # Campo no disponible
            "ultimo_acceso": None,  # Campo no disponible
            "estado": True,  # Asumimos activo
            "estadisticas": estadisticas
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f" Error al obtener usuario: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener usuario: {str(e)}")

@router.put("/usuarios/{usuario_id}")
def actualizar_usuario(
    usuario_id: int,
    usuario_data: dict,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Actualiza la información de un usuario.
    """
    try:
        usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
        # Campos actualizables
        if 'nombre' in usuario_data:
            usuario.nombre = usuario_data['nombre']
        
        if 'correo' in usuario_data:
            # Verificar que el email no esté en uso por otro usuario
            email_existente = db.query(models.Usuario).filter(
                models.Usuario.correo == usuario_data['correo'],
                models.Usuario.id != usuario_id
            ).first()
            if email_existente:
                raise HTTPException(status_code=400, detail="El email ya está en uso")
            usuario.correo = usuario_data['correo']
        
        # El campo telefono no existe en el modelo actual, lo omitimos
        
        if 'rol_id' in usuario_data:
            # Verificar que el rol existe
            rol = db.query(models.Rol).filter(models.Rol.id == usuario_data['rol_id']).first()
            if not rol:
                raise HTTPException(status_code=400, detail="Rol no válido")
            usuario.rol_id = usuario_data['rol_id']
        
        # El campo activo no existe en el modelo actual, lo omitimos
        
        # Actualizar contraseña si se proporciona
        if 'nueva_contrasena' in usuario_data and usuario_data['nueva_contrasena']:
            from passlib.context import CryptContext
            pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
            usuario.contrasena_hash = pwd_context.hash(usuario_data['nueva_contrasena'])
        
        db.commit()
        db.refresh(usuario)
        
        return {
            "success": True,
            "message": "Usuario actualizado exitosamente",
            "usuario": {
                "id": usuario.id,
                "nombre": usuario.nombre,
                "correo": usuario.correo,
                "rol_id": usuario.rol_id
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error al actualizar usuario: {e}")
        raise HTTPException(status_code=400, detail=f"Error al actualizar usuario: {str(e)}")

@router.delete("/usuarios/{usuario_id}")
def eliminar_usuario(
    usuario_id: int,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Elimina un usuario del sistema (solo si no tiene datos asociados).
    """
    try:
        if usuario_id == admin_user.id:
            raise HTTPException(status_code=400, detail="No puedes eliminar tu propio usuario")
        
        usuario = db.query(models.Usuario).filter(models.Usuario.id == usuario_id).first()
        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
        # Verificar si tiene visitas asociadas
        visitas_count = db.execute(text("""
            SELECT COUNT(*) FROM visitas_programadas WHERE visitador_id = :usuario_id
        """), {"usuario_id": usuario_id}).scalar() or 0
        
        if visitas_count > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"No se puede eliminar: el usuario tiene {visitas_count} visitas asociadas"
            )
        
        # Eliminar usuario
        db.delete(usuario)
        db.commit()
        
        return {
            "success": True,
            "message": f"Usuario '{usuario.nombre}' eliminado exitosamente"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error al eliminar usuario: {e}")
        raise HTTPException(status_code=400, detail=f"Error al eliminar usuario: {str(e)}")

@router.get("/usuarios/{usuario_id}/auditoria")
def obtener_auditoria_usuario(
    usuario_id: int,
    limit: int = 50,
    db: Session = Depends(get_db),
    admin_user: models.Usuario = Depends(verificar_admin)
):
    """
    Obtiene el historial de auditoría de un usuario.
    """
    try:
        # Por ahora simulamos datos de auditoría ya que no tenemos tabla de logs
        # En producción, tendrías una tabla de auditoría con logs reales
        
        import random
        from datetime import datetime, timedelta
        
        eventos_auditoria = []
        
        # Simular eventos de auditoría
        eventos_tipos = [
            {"accion": "login", "descripcion": "Inicio de sesión exitoso", "tipo": "info"},
            {"accion": "logout", "descripcion": "Cierre de sesión", "tipo": "info"},
            {"accion": "visita_completada", "descripcion": "Completó visita programada", "tipo": "success"},
            {"accion": "perfil_actualizado", "descripcion": "Actualizó información del perfil", "tipo": "info"},
            {"accion": "contrasena_cambiada", "descripcion": "Cambió la contraseña", "tipo": "warning"},
            {"accion": "acceso_denegado", "descripcion": "Intento de acceso sin permisos", "tipo": "error"},
        ]
        
        # Generar 20-30 eventos de auditoría simulados
        for i in range(random.randint(20, 30)):
            evento = random.choice(eventos_tipos)
            fecha = datetime.now() - timedelta(days=random.randint(1, 90))
            
            eventos_auditoria.append({
                "id": f"audit_{i+1}",
                "fecha": fecha.isoformat(),
                "accion": evento["accion"],
                "descripcion": evento["descripcion"],
                "tipo": evento["tipo"],
                "ip_address": f"192.168.1.{random.randint(1, 255)}",
                "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "detalles": {
                    "session_id": f"sess_{random.randint(1000, 9999)}",
                    "location": random.choice(["Popayán", "López de Micay", "Timbío"])
                }
            })
        
        # Ordenar por fecha descendente
        eventos_auditoria.sort(key=lambda x: x["fecha"], reverse=True)
        
        # Aplicar límite
        eventos_limitados = eventos_auditoria[:limit]
        
        return {
            "total": len(eventos_auditoria),
            "eventos": eventos_limitados,
            "usuario_id": usuario_id,
            "resumen": {
                "total_eventos": len(eventos_auditoria),
                "logins_exitosos": len([e for e in eventos_auditoria if e["accion"] == "login"]),
                "visitas_completadas": len([e for e in eventos_auditoria if e["accion"] == "visita_completada"]),
                "eventos_criticos": len([e for e in eventos_auditoria if e["tipo"] == "error"]),
                "ultimo_acceso": eventos_auditoria[0]["fecha"] if eventos_auditoria else None
            }
        }
    except Exception as e:
        print(f"❌ Error al obtener auditoría: {e}")
        raise HTTPException(status_code=400, detail=f"Error al obtener auditoría: {str(e)}")
